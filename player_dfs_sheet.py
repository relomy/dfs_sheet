import csv
import json
import re
import requests
from os import path
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter, column_index_from_string

from player import Player, QB, RB, WR, TE, DST


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        lcell = row[0]
        rcell = row[-1]
        lcell.border = lcell.border + left
        rcell.border = rcell.border + right
        if fill:
            for c in row:
                c.fill = fill


def create_sheet_header(wb, title, header):
    """Create a sheet within a workbook given a title and header."""
    wb.create_sheet(title=title)
    wb[title].append(header)


def pull_soup_data(filename, ENDPOINT):
    """Either pull file from html or from file."""
    soup = None
    if not path.isfile(filename):
        print("{} does not exist. Pulling from endpoint [{}]".format(filename, ENDPOINT))
        # send GET request
        r = requests.get(ENDPOINT)
        status = r.status_code

        # if not successful, raise an exception
        if status != 200:
            raise Exception('Requests status != 200. It is: {0}'.format(status))

        # dump html to file to avoid multiple requests
        with open(filename, 'w') as outfile:
            print(r.text, file=outfile)

        soup = BeautifulSoup(r.text, 'html5lib')
    else:
        print("File exists [{}]. Nice!".format(filename))
        # load html from file
        with open(filename, 'r') as html_file:
            soup = BeautifulSoup(html_file, 'html5lib')

    return soup


def pull_data(filename, ENDPOINT):
    """Either pull file from API or from file."""
    data = None
    if not path.isfile(filename):
        print("{} does not exist. Pulling from endpoint [{}]".format(filename, ENDPOINT))
        # send GET request
        r = requests.get(ENDPOINT)
        status = r.status_code

        # if not successful, raise an exception
        if status != 200:
            raise Exception('Requests status != 200. It is: {0}'.format(status))

        # store response
        data = r.json()

        # dump json to file for future use to avoid multiple API pulls
        with open(filename, 'w') as outfile:
            json.dump(data, outfile)
    else:
        print("File exists [{}]. Nice!".format(filename))
        # load json from file
        with open(filename, 'r') as json_file:
            data = json.load(json_file)

    return data


def massage_name(name):
    """Remove periods, third names, and special fixes for player names."""
    # remove periods from name
    name = name.replace('.', '')
    # remove Jr. and III etc
    name = ' '.join(name.split(' ')[:2])
    # special fix for Juju Smith-Schuster
    name = name.replace('Juju', 'JuJu')
    return name


def get_fpros_ecr(position):
    """Get stats from FantasyPros for each position."""
    if position == 'QB' or position == 'DST':
        ENDPOINT = 'https://www.fantasypros.com/nfl/rankings/{}.php'.format(
            position.lower())
    else:
        ENDPOINT = 'https://www.fantasypros.com/nfl/rankings/ppr-{}.php'.format(
            position.lower())

    fn = 'ecr_{}.html'.format(position)
    dir = 'sources'
    filename = path.join(dir, fn)

    # pull data
    soup = pull_soup_data(filename, ENDPOINT)

    # find all tables (2) in the html
    table = soup.find('table', id='rank-data')

    if table:
        ls = []

        # # find header
        table_header = table.find('thead')
        # there is one header row
        header_row = table_header.find('tr')
        # loop through header columns and append to worksheet
        header_cols = header_row.find_all('th')
        header = [ele.text.strip() for ele in header_cols]
        ls.append(header)

        # ignore notes
        header = header[:-1]

        # find the rest of the table header_rows
        rows = table.find_all('tr')
        for row in rows:
            cols = row.find_all('td')
            # cols = [ele.text.strip() for ele in cols]
            # change from list comp for just fpros
            new_cols = []
            for ele in cols:
                txt = ele.text.strip()
                # replace JAX
                txt = txt.replace('JAC', 'JAX')
                # remove periods (T.J. Yeldon, T.Y. Hilton)
                txt = txt.replace('.', '')
                # really? just to fix mitchell tribuski?
                if position == 'QB':
                    txt = txt.replace('Mitch', 'Mitchell')
                new_cols.append(txt)

            if cols:
                ls.append(new_cols)

        # return dict(zip(header, new_cols))
        return ls


def get_lineups_player_stats():
    """Meta function to pull all player stats from lineups.com."""
    stats = {
        'snaps': get_lineups_nfl_snaps(),
        'targets': get_lineups_nfl_targets(),
        'receptions': get_lineups_nfl_receptions(),
        'rush_atts': get_lineups_nfl_rush_atts(),
        'redzone_rushes': get_lineups_nfl_redzone_rush_atts(),
        'redzone_targets': get_lineups_nfl_redzone_targets(),
    }
    return stats


def get_lineups_nfl_snaps():
    """Get players' snaps from lineups.com."""
    ENDPOINT = 'https://api.lineups.com/nfl/fetch/snaps/2018/OFF'
    fn = 'nfl_snaps.json'
    dir = 'sources'
    filename = path.join(dir, fn)

    # if file doesn't exist, let's pull it. otherwise - use the file.
    data = pull_data(filename, ENDPOINT)

    if data is None:
        raise Exception('Failed to pull data from API or file.')

    # create dictionary and set key to player's full name
    # dictionary = {}
    # for player in data['data']:
    #     dictionary[player['full_name']] = player

    # return dict comprehension for code above
    return {massage_name(x['full_name']): x for x in data['data']}


def get_lineups_nfl_targets():
    """Get players' targets from lineups.com."""
    ENDPOINT = 'https://api.lineups.com/nfl/fetch/targets/2018/OFF'
    fn = 'nfl_targets.json'
    dir = 'sources'
    filename = path.join(dir, fn)

    # if file doesn't exist, let's pull it. otherwise - use the file.
    data = pull_data(filename, ENDPOINT)

    if data is None:
        raise Exception('Failed to pull data from API or file.')

    # create dictionary and set key to player's full name
    return {massage_name(x['full_name']): x for x in data['data']}


def get_lineups_nfl_receptions():
    """Get players' receptions from lineups.com."""
    ENDPOINT = 'https://api.lineups.com/nfl/fetch/receptions/2018/OFF'
    fn = 'nfl_receptions.json'
    dir = 'sources'
    filename = path.join(dir, fn)

    # if file doesn't exist, let's pull it. otherwise - use the file.
    data = pull_data(filename, ENDPOINT)

    if data is None:
        raise Exception('Failed to pull data from API or file.')

    # create dictionary and set key to player's full name
    return {massage_name(x['name']): x for x in data['data']}


def get_lineups_nfl_rush_atts():
    """Get players' rush attempts from lineups.com."""
    ENDPOINT = 'https://api.lineups.com/nfl/fetch/rush/2018/OFF'
    fn = 'nfl_rush_atts.json'
    dir = 'sources'
    filename = path.join(dir, fn)

    # if file doesn't exist, let's pull it. otherwise - use the file.
    data = pull_data(filename, ENDPOINT)

    if data is None:
        raise Exception('Failed to pull data from API or file.')

    # create dictionary and set key to player's full name
    return {massage_name(x['name']): x for x in data['data']}


def get_lineups_nfl_redzone_rush_atts():
    """Get players' red zone rush attempts from lineups.com."""
    ENDPOINT = 'https://api.lineups.com/nfl/fetch/redzone-rush/2018/OFF'
    fn = 'nfl_redzone_rushes.json'
    dir = 'sources'
    filename = path.join(dir, fn)

    # if file doesn't exist, let's pull it. otherwise - use the file.
    data = pull_data(filename, ENDPOINT)

    if data is None:
        raise Exception('Failed to pull data from API or file.')

    # create dictionary and set key to player's full name
    return {massage_name(x['name']): x for x in data['data']}


def get_lineups_nfl_redzone_targets():
    """Get players' snaps information from lineups.com."""
    red_zone_targets = {}
    for position in ['RB', 'WR', 'TE']:
        ENDPOINT = "https://api.lineups.com/nfl/fetch/redzone-targets/2018/{}".format(
            position)
        fn = "nfl_redzone_targets_{}.json".format(position)
        dir = 'sources'
        filename = path.join(dir, fn)

        # if file doesn't exist, let's pull it. otherwise - use the file.
        data = pull_data(filename, ENDPOINT)

        if data is None:
            raise Exception('Failed to pull data from API or file.')

        red_zone_targets.update({massage_name(x['full_name']): x for x in data['data']})
    # create dictionary and set key to player's full name
    return red_zone_targets


def get_nfl_def_stats(wb):
    """Get teams' defensive stats from lineups.com."""
    # https://www.lineups.com/nfl/teams/stats/defense-stats
    # get passing yds/att
    # td / att (td %)
    # att / completion (compl %)
    ENDPOINT = 'https://api.lineups.com/nfl/fetch/teams/stats/defense-stats/current'
    fn = 'nfl_def_stats.json'
    dir = 'sources'
    filename = path.join(dir, fn)

    # if file doesn't exist, let's pull it. otherwise - use the file.
    data = pull_data(filename, ENDPOINT)

    # we just want player data
    player_data = data['data']

    header = ['team abbv', 'team', 'pass_att', 'pass_yd_per_att', 'pass_compls', 'pass_yd_per_compl',
              'pass_yds', 'pass_tds', 'compl_perc', 'pass_td_per_att_perc']

    team_map = {
        'Atlanta Falcons': 'ATL',
        'Indianapolis Colts': 'IND',
        'San Francisco 49ers': 'SF',
        'Oakland Raiders': 'OAK',
        'Tampa Bay Buccaneers': 'TB',
        'Kansas City Chiefs': 'KC',
        'New York Giants': 'NYG',
        'Cincinnati Bengals': 'CIN',
        'Pittsburgh Steelers': 'PIT',
        'Denver Broncos': 'DEN',
        'Cleveland Browns': 'CLE',
        'New England Patriots': 'NE',
        'Minnesota Vikings': 'MIN',
        'Miami Dolphins': 'MIA',
        'Green Bay Packers': 'GB',
        'Los Angeles Chargers': 'LAC',
        'New Orleans Saints': 'NO',
        'New York Jets': 'NYJ',
        'Arizona Cardinals': 'ARI',
        'Buffalo Bills': 'BUF',
        'Houston Texans': 'HOU',
        'Detroit Lions': 'DET',
        'Jacksonville Jaguars': 'JAX',
        'Los Angeles Rams': 'LAR',
        'Seattle Seahawks': 'SEA',
        'Philadelphia Eagles': 'PHI',
        'Carolina Panthers': 'CAR',
        'Tennessee Titans': 'TEN',
        'Washington Redskins': 'WAS',
        'Dallas Cowboys': 'DAL',
        'Chicago Bears': 'CHI',
        'Baltimore Ravens': 'BAL'
    }
    dictionary = {}
    for d in player_data:
        # TODO rushing_attempt_percentage_by_week
        team = d['team']
        team_abbv = team_map[team]
        pass_att = d['passing_attempts']
        pass_yd_per_att = d['passing_yards_per_attempt']
        pass_compls = d['passing_completions']
        pass_yd_per_compl = d['passing_yards_per_completion']
        pass_yds = d['passing_yards']
        pass_tds = d['passing_touchdowns']

        # personal
        pass_td_per_att_perc = "{0:.4f}".format(pass_tds / pass_att)
        compl_perc = "{0:.4f}".format(pass_compls / pass_att)

        # remove '.' from name
        # name = name.replace('.', '')

        ls = [team_abbv, team, pass_att, pass_yd_per_att, pass_compls, pass_yd_per_compl,
              pass_yds, pass_tds, compl_perc, pass_td_per_att_perc]

        dictionary[team_abbv] = dict(zip(header, ls))
    return dictionary


def conv_weeks_to_padded_list(weeks):
    """Convert weeks dict or list to padded list (16 weeks)."""
    all_weeks = []
    if isinstance(weeks, list):
        for week in weeks:
            if week is None:
                all_weeks.append('')
            else:
                all_weeks.append(week)
    elif isinstance(weeks, dict):
        for i in range(0, len(weeks)):
            # if weeks is None, put in blank string
            # 0 would mean they played but didn't get a snap
            if weeks[str(i + 1)] is None:
                all_weeks.append('')
            else:
                all_weeks.append(weeks[str(i + 1)])

    # pad weeks to 16 (a = [])
    # more visual/pythonic
    # a = (a + N * [''])[:N]
    N = 16
    all_weeks = (all_weeks + N * [''])[:N]
    return all_weeks


def get_vegas_rg(wb):
    """Pull Vegas totals/lines/spreads from RotoGrinders."""
    ENDPOINT = 'https://rotogrinders.com/schedules/nfl'
    fn = 'vegas_script.html'
    dir = 'sources'
    filename = path.join(dir, fn)

    # pull data
    soup = pull_soup_data(filename, ENDPOINT)

    # find script(s) in the html
    script = soup.findAll('script')

    js_vegas_data = script[11].string

    # replace two-letter abbvs
    js_vegas_data = js_vegas_data.replace('GBP', 'GB')
    js_vegas_data = js_vegas_data.replace('JAC', 'JAX')
    js_vegas_data = js_vegas_data.replace('KCC', 'KC')
    js_vegas_data = js_vegas_data.replace('NEP', 'NE')
    js_vegas_data = js_vegas_data.replace('NOS', 'NO')
    js_vegas_data = js_vegas_data.replace('SFO', 'SF')
    js_vegas_data = js_vegas_data.replace('TBB', 'TB')

    # pull json object from data variable
    pattern = re.compile(r'data = (.*);')
    json_str = pattern.search(js_vegas_data).group(1)
    vegas_json = json.loads(json_str)

    vegas = {}
    # iterate through json
    for matchup in vegas_json:
        vegas[matchup['team']] = {
            'display_time': matchup['time']['display'],
            'opponent': matchup['opponent'],
            'line': matchup['line'],
            'moneyline': matchup['moneyline'],
            'overunder': matchup['overunder'],
            'projected': matchup['projected'],
            'projectedchange': matchup['projectedchange']['value']
        }
    return vegas


def get_dvoa_rankings(wb):
    """Get DVOA rankings for team defenses from FootballOutsiders.

    There are two additional get_dvoa_team functions for the resulting tables.
    """
    ENDPOINT = 'https://www.footballoutsiders.com/stats/teamdef'
    fn = 'html_defense.html'
    dir = 'sources'
    filename = path.join(dir, fn)

    # pull data
    soup = pull_soup_data(filename, ENDPOINT)

    # find all tables (3) in the html
    table = soup.findAll('table')

    if table:
        dict_team_rankings = get_dvoa_team_rankings(wb, table[0])
        # separate function for second table
        dict_dvoa_rankings_all = get_dvoa_recv_rankings(
            wb, table[1], dict_team_rankings)

        return dict_dvoa_rankings_all


def get_dvoa_team_rankings(wb, soup_table):
    """Get rankings from first table in HTML."""
    defense_stats = soup_table

    # find header
    # table_header = defense_stats.find('thead')
    # there is one header row
    # header_row = table_header.find('tr')
    # loop through header columns and append to worksheet
    # header_cols = header_row.find_all('th')
    # header = [ele.text.strip() for ele in header_cols]

    # find the rest of the table header_rows
    rows = defense_stats.find_all('tr')

    dvoa_team_rankings = {}

    # make blank NFL key
    dvoa_team_rankings['NFL'] = {}
    for row in rows:
        cols = row.find_all('td')
        cols = [ele.text.strip() for ele in cols]

        if cols:
            # pop 'team_abbv' for dict key
            key = cols.pop(1)

            # na = non-adjusted
            key_names = ['row_num', 'defense_dvoa', 'last_week',
                         'defense_dave', 'total_def_rank', 'pass_def', 'pass_def_rank', 'rush_def',
                         'rush_def_rank', 'na_total', 'na_pass', 'na_rush', 'var', 'sched', 'rank']
            # print(key)
            # map key_names to cols
            dvoa_team_rankings[key] = dict(zip(key_names, cols))
    return dvoa_team_rankings


def get_dvoa_recv_rankings(wb, soup_table, dict_team_rankings):
    """Get rankings from second table in HTML (vs. WR1, WR2, etc.)."""
    # VS types of receivers
    def_recv_stats = soup_table
    # table_header = def_recv_stats.find('thead')
    # header_rows = table_header.find_all('tr')

    rows = def_recv_stats.find_all('tr')
    for row in rows:
        cols = row.find_all('td')
        cols = [ele.text.strip() for ele in cols]

        if cols:
            # pop 'team_abbv' for dict key
            key = cols.pop(0)

            key_names = ['wr1_dvoa', 'wr1_rank', 'wr1_pa_g', 'wr1_yd_g',
                         'wr2_dvoa', 'wr2_rank', 'wr2_pa_g', 'wr2_yd_g',
                         'wro_dvoa', 'wro_rank', 'wro_pa_g', 'wro_yd_g',
                         'te_dvoa', 'te_rank', 'te_pa_g', 'te_yd_g',
                         'rb_dvoa', 'rb_rank', 'rb_pa_g', 'rb_yd_g']
            # print(key)

            # map key_names to cols
            dict_team_rankings[key].update(dict(zip(key_names, cols)))

    return dict_team_rankings


def get_line_rankings(wb):
    """Get offensive and defensive line rankings from FootballOutsiders."""
    dir = 'sources'
    # create empty dict to return
    dictionary = {}
    # dictionary['dl'] = {}
    for line in ['ol', 'dl']:
        ENDPOINT = "https://www.footballoutsiders.com/stats/{0}".format(line)
        fn = "html_{0}.html".format(line)
        filename = path.join(dir, fn)

        # create dict for run and pass stats
        dictionary[line] = {}
        dictionary[line]['run'] = {}
        dictionary[line]['pass'] = {}
        # print(dictionary[line]['run'])

        # pull data
        soup = pull_soup_data(filename, ENDPOINT)

        # find all tables (2) in the html
        table = soup.findAll('table')

        if table:
            # store table
            line_stats = table[0]
            # find header
            # table_header = line_stats.find('thead')
            # there is one header row
            # header_row = table_header.find('tr')
            # loop through header columns and append to worksheet
            # header_cols = header_row.find_all('th')
            # header = [ele.text.strip() for ele in header_cols]

            # find the rest of the table header_rows
            rows = line_stats.find_all('tr')
            for row in rows:
                cols = row.find_all('td')
                cols = [ele.text.strip() for ele in cols]
                if cols:

                    # pop 'team_abbv' for dict key
                    run_key = cols.pop(1)

                    # pop 'team_abbv' for pass protection
                    # pass_key = cols.pop(11)

                    run_key_names = ['rank', 'adj_line_yds', 'rb_yds', 'power_succ_perc', 'power_rank',
                                     'stuff_perc', 'stuff_rank', '2nd_lvl_yds', '2nd_lvl_rank',
                                     'open_field_yds', 'open_field_rank']

                    pass_key_names = ['rank', 'sacks', 'adj_sack_rate']
                    # print(key)

                    # map o-line to 'run' key
                    dictionary[line]['run'][run_key] = dict(zip(run_key_names, cols))
                    # map d-line to 'pass' key
                    # dictionary[line]['pass'][pass_key] = dict(
                    # zip(pass_key_names, cols[-3:]))
                    dictionary[line]['pass'][run_key] = dict(
                        zip(pass_key_names, cols[-3:]))
                    # example
                    # dictionary['ol']['run']['LAR']['adj_line_yds'] = 6.969
                    # dictionary['dl']['pass']['MIA']['adj_sack_rate'] = 4.5%
    return dictionary


def get_matchup_info(game_info, team_abbv):
    """Parse game info into a nice string (home vs. away)."""
    # split game info into matchup_info
    home_team, away_team = game_info.split(' ', 1)[0].split('@')
    if team_abbv == home_team:
        matchup_info = "vs. {}".format(away_team)
    else:
        matchup_info = "at {}".format(home_team)
    return matchup_info


def qb_map(key):
    """Return proper name for dumb abbreviated name from FootballOutsiders."""
    fo_qb_names = {
        'D.Brees': 'Drew Brees',
        'P.Mahomes': 'Patrick Mahomes',
        'J.Goff': 'Jared Goff',
        'P.Rivers': 'Phillip Rivers',
        'M.Ryan': 'Matt Ryan',
        'R.Fitzpatrick': 'Ryan Fitzpatrick',
        'A.Dalton': 'Andy Dalton',
        'J.Flacco': 'Joe Flacco',
        'A.Rodgers': 'Aaron Rodgers',
        'B.Roethlisberger': 'Ben Roethlisberger',
        'K.Cousins': 'Kirk Cousins',
        'T.Brady': 'Tom Brady',
        'D.Carr': 'Derek Carr',
        'M.Trubisky': 'Mitchell Trubisky',
        'D.Watson': 'Deshaun Watson',
        'C.Newton': 'Cam Newton',
        'C.Wentz': 'Carson Wentz',
        'R.Wilson': 'Russell Wilson',
        'J.Winston': 'Jameis Winston',
        'M.Stafford': 'Matthew Stafford',
        'S.Darnold': 'Sam Darnold',
        'A.Luck': 'Andrew Luck',
        'C.Keenum': 'Case Keenum',
        'C.J.Beathard': 'CJ Beathard',
        'A.Smith': 'Alex Smith',
        'J.Rosen': 'Josh Rosen',
        'B.Bortles': 'Blake Bortles',
        'E.Manning': 'Eli Manning',
        'J.Garoppolo': 'Jimmy Garoppolo',
        'R.Tannehill': 'Ryan Tannehill',
        'D.Prescott': 'Dak Prescott',
        'M.Mariota': 'Marcus Mariota',
        'B.Mayfield': 'Baker Mayfield',
        'T.Taylor': 'Tyrod Taylor',
        'J.Allen': 'Josh Allen',
        'B.Osweiler': 'Brock Osweiler',
        'B.Gabbert': 'Blaine Gabbert',
        'C.Kessley': 'Cody Kessler',
        'D.Enderson': 'Derek Anderson',
        'N.Foles': 'Nick Foles',
        'S.Bradford': 'Sam Bradford',
        'N.Peterman': 'Nathan Peterman',
        'T.Taylor': 'Tyrod Taylor',
        'L.Jackson': 'Lamar Jackson'
    }
    return fo_qb_names.get(key, None)


def get_qb_stats_FO(wb):
    """Get QB stats from FootballOutsidersself.

    There are three separate tables that need to be parsed.
    """
    ENDPOINT = 'https://www.footballoutsiders.com/stats/qb'
    fn = 'html_qb.html'
    dir = 'sources'
    filename = path.join(dir, fn)

    # pull data
    soup = pull_soup_data(filename, ENDPOINT)

    # find all tables (3) in the html
    table = soup.findAll('table')

    if table:
        dictionary = {}
        for i, t in enumerate(table):
            # find header
            table_header = t.find('thead')
            # there is one header row
            header_row = table_header.find('tr')
            # loop through header columns and append to worksheet
            header_cols = header_row.find_all('th')
            header = [ele.text.strip() for ele in header_cols]

            # find the rest of the table header_rows
            rows = t.find_all('tr')
            for row in rows:
                cols = row.find_all('td')
                cols = [ele.text.strip().replace(',', '') for ele in cols]
                if cols:
                    # pop 'name' for dict key
                    key = cols.pop(0)

                    # i only create this list because Lamar Jackson has no pass_dyar but he has rushing stats
                    # also Nathan Peterman has no rushing yds..
                    main_fields = ['team', 'pass_dyar', 'dyar_rank', 'yar', 'yar_rank', 'pass_dvoa', 'pass_dvoa_rank', 'voa',
                                   'qbr', 'qbr_rank', 'pass_atts', 'pass_yds', 'eyds', 'tds', 'fk', 'fl', 'int', 'c_perc', 'dpi', 'alex',
                                   'rush_dyar', 'rush_dyar_rank', 'rush_yar', 'rush_yar_rank', 'rush_dvoa', 'rush_dvoa_rank', 'rush_voa', 'rush_atts', 'rush_yds', 'rush_eyds', 'rush_tds', 'fumbles']
                    if i == 0:
                        key_names = ['team', 'pass_dyar', 'dyar_rank', 'yar', 'yar_rank', 'pass_dvoa', 'pass_dvoa_rank',
                                     'voa', 'qbr', 'qbr_rank', 'pass_atts', 'pass_yds', 'eyds', 'tds', 'fk', 'fl',
                                     'int', 'c_perc', 'dpi', 'alex']
                    elif i == 1:
                        key_names = ['team', 'pass_dyar', 'pass_dvoa', 'pass_dvoa_rank', 'voa', 'qbr',
                                     'qbr_rank', 'pass_atts', 'pass_yds', 'eyds', 'tds', 'fk', 'fl', 'int', 'c_perc', 'dpi', 'alex']
                    elif i == 2:
                        key_names = ['team', 'rush_dyar', 'rush_dyar_rank', 'rush_yar', 'rush_yar_rank', 'rush_dvoa', 'rush_dvoa_rank',
                                     'rush_voa', 'rush_atts', 'rush_yds', 'rush_eyds', 'rush_tds', 'fumbles']

                    # print(key)
                    # map key_names to cols

                    player_name = qb_map(key)
                    # create dictionary if it does not exist
                    if player_name not in dictionary:
                        dictionary[player_name] = dict.fromkeys(main_fields, None)
                    dictionary[player_name].update(dict(zip(key_names, cols)))
    return dictionary


def find_name_in_ecr(ecr_pos_list, name):
    for item in ecr_pos_list:
        # if any(name in s for s in item):
        # look in 2nd column for name
        if len(item) > 2 and name in item[2]:
            # print("Found {}!".format(name))
            # rank, wsis, dumb_name, matchup, best, worse, avg, std_dev = item
            return item
    return False


def read_fantasy_draft_csv(filename):
    team_map = {
        'Atlanta Falcons': 'ATL',
        'Indianapolis Colts': 'IND',
        'San Francisco 49ers': 'SF',
        'Oakland Raiders': 'OAK',
        'Tampa Bay Buccaneers': 'TB',
        'Kansas City Chiefs': 'KC',
        'New York Giants': 'NYG',
        'Cincinnati Bengals': 'CIN',
        'Pittsburgh Steelers': 'PIT',
        'Denver Broncos': 'DEN',
        'Cleveland Browns': 'CLE',
        'New England Patriots': 'NE',
        'Minnesota Vikings': 'MIN',
        'Miami Dolphins': 'MIA',
        'Green Bay Packers': 'GB',
        'Los Angeles Chargers': 'LAC',
        'New Orleans Saints': 'NO',
        'New York Jets': 'NYJ',
        'Arizona Cardinals': 'ARI',
        'Buffalo Bills': 'BUF',
        'Houston Texans': 'HOU',
        'Detroit Lions': 'DET',
        'Jacksonville Jaguars': 'JAX',
        'Los Angeles Rams': 'LAR',
        'Seattle Seahawks': 'SEA',
        'Philadelphia Eagles': 'PHI',
        'Carolina Panthers': 'CAR',
        'Tennessee Titans': 'TEN',
        'Washington Redskins': 'WAS',
        'Dallas Cowboys': 'DAL',
        'Chicago Bears': 'CHI',
        'Baltimore Ravens': 'BAL'
    }

    with open(filename, 'r') as f:
        reader = csv.reader(f)

        # store header row (and strip extra spaces)
        headers = [header.lower().strip() for header in next(reader)]
        headers.append('salary_perc')

        # fill dictionary to return
        dictionary = {}
        for row in reader:
            if row[0] == 'DST':
                # map full team name to team abbv
                row[1] = team_map.get(row[1], None)
            # remove periods from name
            row[1] = row[1].replace('.', '')
            # remove Jr. and III etc
            row[1] = ' '.join(row[1].split(' ')[:2])
            # store salary without $ or ,
            row[5] = row[5][1:].replace(',', '')
            # calculate salary percentage
            salary_perc = "{0:0.1%}".format(float(row[5]) / 100000)
            row.append(salary_perc)
            dictionary[row[1]] = {key: value for key, value in zip(headers, row)}
        return dictionary


def excel_write_position_to_sheet(wb, player):
    # create
    if player.position not in wb.sheetnames:
        wb.create_sheet(title=player.position)
        # create top level header for positional tab
        excel_write_top_level_header(wb[player.position], player)
        wb[player.position].append(player.get_writable_header())

    ws = wb[player.position]
    # print("max_row: {}".format(ws.max_row))
    ws.append(player.get_writable_row())
    excel_apply_format_row(ws, ws.max_row)


def excel_apply_format_row(ws, row_num):
    al = Alignment(horizontal='center', vertical='center')
    for cell in ws[row_num]:
        cell.alignment = al


def excel_insert_ranks(wb):
    """In each positional tab, find columns from header and fill in ranks."""
    header_row_num = 2
    for position in ['QB', 'RB', 'WR', 'TE', 'DST']:
        ws = wb[position]

        ecr_col = ''
        ecr_data_col = ''
        salary_col = ''
        salary_rank_col = ''
        fd_salary_col = ''
        fd_salary_rank_col = ''
        plus_minus_col = ''
        fd_plus_minus_col = ''
        max_row = ws.max_row

        # look through header row and pull header columns
        for col in ws[header_row_num]:
            if col.value == 'ECR':
                ecr_col = col.column
            elif col.value == 'ECR Data':
                ecr_data_col = col.column
            elif col.value == 'Salary':
                salary_col = col.column
            elif col.value == 'Salary Rank':
                salary_rank_col = col.column
            elif col.value == '+/- Rank':
                plus_minus_col = col.column
            elif col.value == 'FD Salary':
                fd_salary_col = col.column
            elif col.value == 'FDraft Salary Rank':
                fd_salary_rank_col = col.column
            elif col.value == 'FD +/- Rank':
                fd_plus_minus_col = col.column

        # ECR rank
        for cell in ws[ecr_col]:
            # skip header rows
            if cell.row <= 2:
                continue
            cell.value = '=RANK(${0}{1}, ${0}3:${0}{2},1)'.format(
                ecr_data_col, cell.row, max_row)

        # salary rank
        for cell in ws[salary_rank_col]:
            # skip header rows
            if cell.row <= 2:
                continue
            cell.value = '=RANK(${0}{1}, ${0}3:${0}{2},0)'.format(
                salary_col, cell.row, max_row)

        # +/- rank
        for cell in ws[plus_minus_col]:
            # skip header rows
            if cell.row <= 2:
                continue
            cell.value = '={0}{1} - {2}{1}'.format(
                salary_rank_col, cell.row, ecr_col)

        # FD salary rank
        for cell in ws[fd_salary_rank_col]:
            # skip header rows
            if cell.row <= 2:
                continue
            cell.value = '=RANK(${0}{1}, ${0}3:${0}{2},0)'.format(
                fd_salary_col, cell.row, max_row)

        # fd +- rank - fd salary rank - DK salary rank
        for cell in ws[fd_plus_minus_col]:
            # skip header rows
            if cell.row <= 2:
                continue
            cell.value = '={0}{1} - {2}{1}'.format(
                fd_salary_rank_col, cell.row, salary_rank_col)
        # hide data columns
        # print("1: {}".format(ecr_data_col))
        # print("2: {}".format(salary_rank_col))
        # print("3: {}".format(ws))
        # print("4: {}".format(ws.column_dimensions[ecr_data_col]))
        # ws.column_dimensions[ecr_data_col].hidden = True
        # ws.column_dimensions[salary_rank_col].hidden = True


def excel_write_top_level_header(ws, player):
    """Write the top most header row with merged cells and colors."""
    # colors
    color_white = 'FF000000'
    color_yellow = 'FFFFC000'
    color_orange = 'FFED7D31'
    color_light_blue = 'FF00B0F0'
    color_darker_blue = 'FF5B9BD5'
    color_green = 'FF70AD47'
    color_cyan = 'FFA8F3D9'

    # these values are in all positional tabs
    dictionary = {
        'DK': {'length': 2, 'color': color_white},
        'VEGAS': {'length': 3, 'color': color_yellow},
        'RANKINGS': {'length': 3, 'color': color_green},
        'FDRAFT': {'length': 3, 'color': color_cyan}
    }
    if player.position == 'QB':
        header_order = ['VEGAS', 'SEASON',
                        'PRESSURE', 'MATCHUP', 'RANKINGS', 'DK', 'FDRAFT']
        dictionary['SEASON'] = {'length': 3, 'color': color_light_blue}
        dictionary['PRESSURE'] = {'length': 2, 'color': color_darker_blue}
        dictionary['MATCHUP'] = {'length': 4, 'color': color_orange}
    elif player.position == 'RB':
        header_order = ['VEGAS', 'MATCHUP',
                        'SEASON', 'LAST WEEK', 'RANKINGS', 'DK', 'FDRAFT']
        dictionary['MATCHUP'] = {'length': 4, 'color': color_orange}
        dictionary['SEASON'] = {'length': 4, 'color': color_light_blue}
        dictionary['LAST WEEK'] = {'length': 4, 'color': color_darker_blue}
    elif player.position == 'WR':
        header_order = ['VEGAS', 'MATCHUP',
                        'SEASON', 'LAST WEEK', 'RANKINGS', 'DK', 'FDRAFT']
        dictionary['MATCHUP'] = {'length': 3, 'color': color_orange}
        dictionary['SEASON'] = {'length': 4, 'color': color_light_blue}
        dictionary['LAST WEEK'] = {'length': 4, 'color': color_darker_blue}

    elif player.position == 'TE':
        header_order = ['VEGAS', 'MATCHUP',
                        'SEASON', 'LAST WEEK', 'RANKINGS', 'DK', 'FDRAFT']
        dictionary['MATCHUP'] = {'length': 2, 'color': color_orange}
        dictionary['SEASON'] = {'length': 4, 'color': color_light_blue}
        dictionary['LAST WEEK'] = {'length': 4, 'color': color_darker_blue}
    elif player.position == 'DST':
        header_order = ['VEGAS', 'RANKINGS', 'DK', 'FDRAFT']

    # start column count after 'blank' header
    column_count = 5
    for header in header_order:
        start_column = get_column_letter(column_count)
        excel_merge_top_header(
            ws, header, start_column, dictionary[header]['length'], dictionary[header]['color'])
        column_count += dictionary[header]['length']


def excel_merge_top_header(ws, text, start_col, length, color):
    """Style a range as merged cells."""
    row_num = 1
    # alignment style for merge + center
    al = Alignment(horizontal="center", vertical="center")
    # bold font
    font = Font(b=True, color="FFFFFFFF")

    # we actually want the length to be inclusive of the start column
    length -= 1

    # get coloumn index from string
    start_col_idx = column_index_from_string(start_col)

    # insert text into cell
    ws.cell(row=1, column=start_col_idx).value = text

    # set range to format merged cells
    fmt_range = "{0}{row_num}:{1}{row_num}".format(
        start_col, get_column_letter(start_col_idx + length), row_num=row_num)
    # fmt_range = "{0}1:{1}1".format(get_column_letter(
    #     start_col), get_column_letter(start_col + length))
    style_range(ws, fmt_range, font=font, fill=PatternFill(
        patternType="solid", fgColor=color), alignment=al)


def excel_apply_borders(wb):
    header_row = 1
    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'))
    bottom_side_border = Border(bottom=Side(border_style='thin', color='FF000000'))

    for position in ['QB', 'RB', 'WR', 'TE', 'DST']:
        # select worksheet
        ws = wb[position]
        # find header columns (None = empty cell)
        fields = [cell.column for cell in ws[header_row] if cell.value is not None]
        # for cell in ws[1]:
        #     if cell.value is not None:
        #         fields.append(cell.column)
        # print("field: {} [{}] [idx: {}]".format(cell.value, cell.column, cell.col_idx))

        # add max column (letter) to field
        # fields.append(get_column_letter(ws.max_column))
        # add arbitrary 3 letters for fantasy draft
        last_column_header_index = column_index_from_string(fields[-1])
        final_column_for_border = get_column_letter(last_column_header_index + 3)
        fields.append(final_column_for_border)

        # skip first field
        for i in range(1, len(fields)):
            prev_letter_index = column_index_from_string(fields[i]) - 1
            prev_letter = get_column_letter(prev_letter_index)
            fmt_range = "{0}1:{1}{2}".format(
                fields[i - 1], prev_letter, ws.max_row)
            style_range(ws, fmt_range, border=border)

        # add bottom border on last row
        for cell in ws[ws.max_row]:
            cell.border = cell.border + bottom_side_border


def find_fields_in_header(ws, search_fields):
    """Search one ws header for many fields. Return indices of header to be numberified."""
    header_row = 2
    columns = []
    for cell in ws[header_row]:
        if cell.value in search_fields:
            columns.append(cell.column)
            continue
    return columns


def excel_apply_format_header(wb):
    header_row_num = 2
    row_height = 40
    # change row font and alignment
    font = Font(b=True, color="FF000000")
    al = Alignment(horizontal="center", vertical="center", wrapText=True)

    for position in ['QB', 'RB', 'WR', 'TE', 'DST']:
        ws = wb[position]
        # set row height
        ws.row_dimensions[header_row_num].height = row_height

        for cell in ws[header_row_num]:
            cell.font = font
            cell.alignment = al


def excel_apply_cell_number_formats(wb):
    percentage_fields = ['Salary%', 'FD Salary%', 'O-Line Sack%', 'D-Line Sack%',
                         'Def Comp%', 'Def TD%']
    currency_fields = ['Salary', 'FD Salary']
    for position in ['QB', 'RB', 'WR', 'TE', 'DST']:
        ws = wb[position]

        # apply percentage format to percent fields
        columns_perc = find_fields_in_header(ws, percentage_fields)
        for column in columns_perc:
            for cell in ws[column]:
                cell.number_format = '##0.0%'

        # apply dollar format to salary fields
        columns_currency = find_fields_in_header(ws, currency_fields)
        for column in columns_currency:
            for cell in ws[column]:
                cell.number_format = '$#,##0_);($#,##0)'


def excel_apply_column_widths(wb):
    """Apply column widths to positional tabs."""
    for position in ['QB', 'RB', 'WR', 'TE', 'DST']:
        ws = wb[position]
        for i, cell in enumerate(ws[2]):
            # print(cell)
            if cell.value == 'Name':
                ws.column_dimensions[get_column_letter(i + 1)].width = 20
            elif cell.value in ['FD Salary', 'Opp']:
                ws.column_dimensions[get_column_letter(i + 1)].width = 9
            elif cell.value in ['Position']:
                ws.column_dimensions[get_column_letter(i + 1)].width = 8
            elif cell.value in ['FD +/- Rank']:
                ws.column_dimensions[get_column_letter(i + 1)].width = 6
            elif cell.value in ['+/- Rank', 'Line', 'ECR']:
                ws.column_dimensions[get_column_letter(i + 1)].width = 5
            else:
                ws.column_dimensions[get_column_letter(i + 1)].width = 7


def excel_apply_conditional_formatting(wb):
    # define colors for colorscale (from excel)
    red = 'F8696B'
    yellow = 'FFEB84'
    green = '63BE7B'
    white = 'FFFFFF'

    # start color rule after headers
    start_row = 3

    for position in ['QB', 'RB', 'WR', 'TE', 'DST']:
        ws = wb[position]
        # add filter/sort. excel will not automatically do it!
        # filter_range = "{0}:{1}".format('D2', ws.max_row)
        # ws.auto_filter.ref = filter_range
        # sort_range = "{0}:{1}".format('D3', ws.max_row)

        # ws.auto_filter.add_sort_condition(sort_range)
        # bigger/positive = green, smaller/negative = red
        green_to_red_headers = [
            'Total', 'O/U', 'Run DVOA', 'Pass DVOA', 'DVOA', 'vs. WR1', 'vs. WR2',
            'O-Line', 'D-Line', 'Snap%', 'Rush ATTs', 'Trgts', 'Rcpts', 'vs. TE',
            'Ave PPG', 'Rush Yards', 'DYAR', 'QBR', 'Def Yds/Att', 'Def Comp%', 'Def TD%',
            'RZ Opps'
        ]
        green_to_red_rule = ColorScaleRule(start_type='min', start_color=red,
                                           mid_type='percentile', mid_value=50, mid_color=yellow,
                                           end_type='max', end_color=green)
        # bigger/positive = red, smaller/negative = green
        red_to_green_headers = [
            'Line', 'O-Line Sack%', 'D-Line Sack%', 'ECR'
        ]
        red_to_green_rule = ColorScaleRule(start_type='min', start_color=green,
                                           mid_type='percentile', mid_value=50, mid_color=yellow,
                                           end_type='max', end_color=red)
        white_middle_headers = [
            '+/- Rank', 'FD +/- Rank'
        ]
        white_middle_rule = ColorScaleRule(start_type='min', start_color=red,
                                           mid_type='percentile', mid_value=50, mid_color=white,
                                           end_type='max', end_color=green)
        # color ranges
        for i in range(1, ws.max_column + 1):
            column_letter = get_column_letter(i)
            cell_rng = "{0}{1}:{2}".format(column_letter, start_row, ws.max_row)
            if ws.cell(row=2, column=i).value in green_to_red_headers:
                # color range (green to red)
                ws.conditional_formatting.add(cell_rng, green_to_red_rule)
            elif ws.cell(row=2, column=i).value in red_to_green_headers:
                # color range (red to green)
                ws.conditional_formatting.add(cell_rng, red_to_green_rule)
            elif ws.cell(row=2, column=i).value in white_middle_headers:
                # color range (red to white to green)
                ws.conditional_formatting.add(cell_rng, white_middle_rule)


def excel_apply_hide_columns(wb):
    header_row_num = 2
    hidden_columns = ['Abbv', 'ECR Data', 'Salary Rank', 'FDraft Salary Rank']
    for position in ['QB', 'RB', 'WR', 'TE', 'DST']:
        ws = wb[position]
        for col in ws[header_row_num]:
            if col.value in hidden_columns:
                ws.column_dimensions[col.column].hidden = True


def excel_apply_header_freeze(wb):
    # freeze header
    row_id = 3
    for position in ['QB', 'RB', 'WR', 'TE', 'DST']:
        ws = wb[position]
        # panes frozen are above and to the left of the cell frozen
        ws.freeze_panes = "D{}".format(row_id)
        # max_row = ws.max_row
        # ws.freeze_panes = "{0}{row_id}".format(
        #     get_column_letter(ws.max_column), row_id=row_id)


def excel_apply_filter_setup(wb):
    """Apply filter to second header row."""
    header_row = 2
    for position in ['QB', 'RB', 'WR', 'TE', 'DST']:
        ws = wb[position]
        filter_rng = "A{0}:{1}{0}".format(header_row, get_column_letter(ws.max_column))
        ws.auto_filter.ref = filter_rng


def excel_apply_sheet_order(wb):
    """Re-order sheet tabs using private variable."""
    # pull indices from QB, RB, WR, TE, DST to be ordered first
    order = [wb.worksheets.index(wb[i]) for i in ['QB', 'RB', 'WR', 'TE', 'DST']]

    # create set from 0 to len(wb._sheets)
    # subtract unique values from set and extend list to fill in missing values
    order.extend(list(set(range(len(wb._sheets))) - set(order)))
    wb._sheets = [wb._sheets[i] for i in order]


def main():
    fn = 'DKSalaries_week9_full.csv'
    dest_filename = 'player_sheet.xlsx'

    # create workbook/worksheet
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'DEL'

    # guess types (numbers, floats, etc)
    wb.guess_types = True

    # dict  of players (key = DFS player name)
    # player_dict = {}

    # make sources dir if it does not exist
    # directory = 'sources'
    # if not path.exists(directory):
    #     makedirs(directory)

    # pull positional stats from fantasypros.com
    ecr_pos_dict = {}
    # for position in ['QB', 'RB', 'WR', 'TE', 'DST']:
    for position in ['QB', 'RB', 'WR', 'TE', 'DST']:
        ecr_pos_dict[position] = get_fpros_ecr(position)

    fdraft_csv = 'FDraft_week9_full.csv'
    if path.exists(fdraft_csv):
        fdraft_dict = read_fantasy_draft_csv(fdraft_csv)
    else:
        fdraft_dict = None

    # vegas lines from rotogrinders.com
    vegas_dict = get_vegas_rg(wb)
    # get snaps, targets, receptions, rush attempts from lineups.com
    stats_dict = get_lineups_player_stats()
    # defense stats from lineups.com
    def_dict = get_nfl_def_stats(wb)
    # DVOA rankings from footballoutsiders.com
    dvoa_dict = get_dvoa_rankings(wb)
    # OL/DL rankings from footballoutsiders.com
    line_dict = get_line_rankings(wb)
    # QB rankings from footballoutsiders.com
    qb_dict = get_qb_stats_FO(wb)

    # print(dvoa_dict['CHI'])

    # create list for players
    player_list = []

    with open(fn, 'r') as f:
        # read entire file into memory
        lines = f.readlines()

        for i, line in enumerate(lines):
            # skip header
            if i == 0:
                continue

            fields = line.rstrip().split(',')

            # store each variable from list
            position, name_id, name, id, roster_pos, salary, game_info, team_abbv, average_ppg = fields

            # 'fix' name to remove extra stuff like Jr or III (Todd Gurley II for example)
            name = ' '.join(name.split(' ')[:2])
            # also remove periods (T.J. Yeldon for example)
            name = name.replace('.', '')

            matchup = get_matchup_info(game_info, team_abbv)
            # use team_abbv for DSTs
            if position == 'DST':
                name = fields[7]

            # if player is not in ECR rankings, skip him
            ecr_item = find_name_in_ecr(ecr_pos_dict[position], name)
            if ecr_item:
                # ecr_rank, ecr_wsis, ecr_dumb_name, ecr_matchup, ecr_best, ecr_worse, ecr_avg, ecr_std_dev = ecr_item
                ecr_rank = ecr_item[0]
                # if position == 'DST':
                #     print(fields)
                #     print("{}".format(ecr_item))
                #     print("Name: {} matchup: {}".format(name, matchup))
                #     print()

                # create Player class for position
                p = Player(name, position, team_abbv, salary, game_info,
                           average_ppg, matchup, ecr_rank)

                if fdraft_dict:
                    p.set_fdraft_fields(fdraft_dict[name]['salary'],
                                        fdraft_dict[name]['salary_perc'])

                # set vegas fields based on team abbv (key)
                p.set_vegas_fields(
                    vegas_dict[team_abbv]['overunder'], vegas_dict[team_abbv]['line'], vegas_dict[team_abbv]['projected'])

                if position == 'QB':
                    qb = QB(p)

                    # convert string ('3.8%') to float (0.038)
                    line_sack_rate = line_dict['ol']['pass'][team_abbv]['adj_sack_rate'].replace(
                        '%', '')
                    opp_sack_rate = line_dict['dl']['pass'][p.opponent]['adj_sack_rate'].replace(
                        '%', '')
                    qb.line_sack_rate = float(line_sack_rate) / 100
                    qb.opp_sack_rate = float(opp_sack_rate) / 100

                    # check for QB in qb_dict
                    if name in qb_dict:
                        qb.rush_yds = qb_dict[name]['rush_yds']
                        qb.pass_dyar = qb_dict[name]['pass_dyar']
                        qb.qbr = qb_dict[name]['qbr']
                    else:
                        print("Could find no QB information on {}".format(name))

                    # check for opponent in def_dict
                    if p.opponent in def_dict:
                        qb.pass_def_rank = dvoa_dict[p.opponent]['pass_def_rank']
                        qb.opp_yds_att = def_dict[p.opponent]['pass_yd_per_att']
                        qb.opp_comp_perc = def_dict[p.opponent]['compl_perc']
                        qb.opp_td_perc = def_dict[p.opponent]['pass_td_per_att_perc']
                    else:
                        print("Could find no DEF information for {}".format(p.opponent))

                    player_list.append(qb)
                elif position == 'RB':
                    rb = RB(p)

                    # set position-specific dvoa fields
                    rb.run_dvoa = dvoa_dict[p.opponent]['rush_def_rank']
                    rb.rb_pass_dvoa = dvoa_dict[p.opponent]['rb_rank']

                    # set oline/opponent dline stats for adjusted line yards
                    rb.oline_adj_line_yds = line_dict['ol']['run'][team_abbv]['adj_line_yds']
                    rb.opp_adj_line_yds = line_dict['dl']['run'][p.opponent]['adj_line_yds']

                    if name in stats_dict['snaps']:
                        # set season numbers
                        rb.season_snap_percent = stats_dict['snaps'][name]['season_snap_percent']
                        rb.season_rush_atts = stats_dict['rush_atts'][name]['average']
                        rb.season_targets = stats_dict['targets'][name]['average']

                        # store lists in Player object
                        rb.snap_percentage_by_week = stats_dict['snaps'][name]['snap_percentage_by_week']
                        rb.rush_atts_weeks = stats_dict['rush_atts'][name]['weeks']
                        rb.targets_weeks = stats_dict['targets'][name]['weeks']

                        # currently need a class method here to calculate/set last weeks snaps/etc stats
                        rb.set_last_week_fields()
                    else:
                        print("Could find no SNAPS information on {} [{}]".format(
                            name, position))

                    # look for redzone opportunities
                    if name in stats_dict['redzone_targets']:
                        rb.season_rz_avg_targets = stats_dict['redzone_targets'][name]['average']
                        rb.rz_targets_weeks = stats_dict['redzone_targets'][name]['weeks']
                    if name in stats_dict['redzone_rushes']:
                        rb.season_rz_avg_rush_atts = stats_dict['redzone_rushes'][name]['average']
                        rb.rz_rush_atts_weeks = stats_dict['redzone_rushes'][name]['weeks']

                    rb.season_rz_opps = rb.season_rz_avg_targets + rb.season_rz_avg_rush_atts

                    # store lists in Player object

                    # currently need a class method here to calculate/set last weeks snaps/etc stats
                    rb.set_last_week_rz_fields()

                    if rb.last_week_rz_rush_atts is None and rb.last_week_rz_targets is None:
                        rb.last_week_rz_opps = 0
                    else:
                        if rb.last_week_rz_rush_atts is None:
                            rb.last_week_rz_rush_atts = 0

                        if rb.last_week_rz_targets is None:
                            rb.last_week_rz_targets = 0
                        rb.last_week_rz_opps = rb.last_week_rz_rush_atts + rb.last_week_rz_targets

                    # print("[{}] rush: {} + targets: {} = opps: {}".format(rb.name,
                    #                                                       rb.last_week_rz_rush_atts,
                    #                                                       rb.last_week_rz_targets,
                    #                                                       rb.last_week_rz_opps))
                    # call class method to set fields for last week
                    player_list.append(rb)
                elif position == 'WR':
                    wr = WR(p)
                    # set position-specific dvoa fields
                    wr.pass_def_rank = dvoa_dict[p.opponent]['pass_def_rank']
                    wr.wr1_rank = dvoa_dict[p.opponent]['wr1_rank']
                    wr.wr2_rank = dvoa_dict[p.opponent]['wr2_rank']

                    # if player is not in snaps, he likely has no other information either
                    if name in stats_dict['snaps']:
                        # set season numbers
                        wr.season_snap_percent = stats_dict['snaps'][name]['season_snap_percent']
                        wr.season_targets = stats_dict['targets'][name]['average']
                        wr.season_recepts = stats_dict['receptions'][name]['average']

                        # store lists in Player object
                        wr.snap_percentage_by_week = stats_dict['snaps'][name]['snap_percentage_by_week']
                        wr.recepts_weeks = stats_dict['receptions'][name]['weeks']
                        wr.targets_weeks = stats_dict['targets'][name]['weeks']

                        # call class method to set fields for last week
                        wr.set_last_week_fields()
                    else:
                        print("Could find no SNAPS information on {} [{}]".format(
                            name, position))

                    # look for redzone opportunities
                    if name in stats_dict['redzone_targets']:
                        wr.season_rz_avg_targets = stats_dict['redzone_targets'][name]['average']
                        wr.rz_targets_weeks = stats_dict['redzone_targets'][name]['weeks']
                    if name in stats_dict['redzone_rushes']:
                        wr.season_rz_avg_rush_atts = stats_dict['redzone_rushes'][name]['average']
                        wr.rz_rush_atts_weeks = stats_dict['redzone_rushes'][name]['weeks']

                    wr.season_rz_opps = wr.season_rz_avg_targets + wr.season_rz_avg_rush_atts

                    # currently need a class method here to calculate/set last weeks snaps/etc stats
                    wr.set_last_week_rz_fields()

                    if wr.last_week_rz_rush_atts is None and wr.last_week_rz_targets is None:
                        wr.last_week_rz_opps = 0
                    else:
                        if wr.last_week_rz_rush_atts is None:
                            wr.last_week_rz_rush_atts = 0

                        if wr.last_week_rz_targets is None:
                            wr.last_week_rz_targets = 0
                        wr.last_week_rz_opps = wr.last_week_rz_rush_atts + wr.last_week_rz_targets

                    # print("[{}] rush: {} + targets: {} = opps: {}".format(wr.name,
                    #                                                       wr.last_week_rz_rush_atts,
                    #                                                       wr.last_week_rz_targets,
                    #                                                       wr.last_week_rz_opps))

                    player_list.append(wr)
                elif position == 'TE':
                    te = TE(p)
                    # set position-specific dvoa fields
                    te.pass_def_rank = dvoa_dict[p.opponent]['pass_def_rank']
                    te.te_rank = dvoa_dict[p.opponent]['te_rank']

                    if name in stats_dict['snaps']:
                        # set season numbers
                        te.season_snap_percent = stats_dict['snaps'][name]['season_snap_percent']
                        te.season_targets = stats_dict['targets'][name]['average']
                        te.season_recepts = stats_dict['receptions'][name]['average']

                        te.snap_percentage_by_week = stats_dict['snaps'][name]['snap_percentage_by_week']
                        te.recepts_weeks = stats_dict['receptions'][name]['weeks']
                        te.targets_weeks = stats_dict['targets'][name]['weeks']

                        # call class method to set fields for last week
                        te.set_last_week_fields()
                        # print("snaps list: {}".format(rb.snap_percentage_by_week))
                        # print(te.snap_percentage_by_week)
                        # print(te.recepts_weeks)
                        # print(te.targets_weeks)
                        # print("last_week_snaps: {}".format(wr.last_week_snaps()))
                        # print("last_week_rush: {}".format(wr.last_week_rush_atts()))
                        # print("last_week_targets: {}".format(wr.last_week_targets()))
                        # exit()
                    else:
                        print("Could find no SNAPS information on {} [{}]".format(
                            name, position))

                    # look for redzone opportunities
                    if name in stats_dict['redzone_targets']:
                        te.season_rz_avg_targets = stats_dict['redzone_targets'][name]['average']
                        te.rz_targets_weeks = stats_dict['redzone_targets'][name]['weeks']
                    if name in stats_dict['redzone_rushes']:
                        te.season_rz_avg_rush_atts = stats_dict['redzone_rushes'][name]['average']
                        te.rz_rush_atts_weeks = stats_dict['redzone_rushes'][name]['weeks']

                    te.season_rz_opps = te.season_rz_avg_targets + te.season_rz_avg_rush_atts

                    # currently need a class method here to calculate/set last weeks snaps/etc stats
                    te.set_last_week_rz_fields()

                    if te.last_week_rz_rush_atts is None and te.last_week_rz_targets is None:
                        te.last_week_rz_opps = 0
                    else:
                        if te.last_week_rz_rush_atts is None:
                            te.last_week_rz_rush_atts = 0

                        if te.last_week_rz_targets is None:
                            te.last_week_rz_targets = 0
                        te.last_week_rz_opps = te.last_week_rz_rush_atts + te.last_week_rz_targets

                    # print("[{}] rush: {} + targets: {} = opps: {}".format(te.name,
                    #                                                       te.last_week_rz_rush_atts,
                    #                                                       te.last_week_rz_targets,
                    #                                                       te.last_week_rz_opps))

                    player_list.append(te)
                elif position == 'DST':
                    dst = DST(p)
                    player_list.append(dst)
            # else:
            #     print("{} not found in ECR rankings".format(name))

    for player in player_list:
        excel_write_position_to_sheet(wb, player)

    # apply Excel functions
    excel_insert_ranks(wb)
    excel_apply_format_header(wb)
    excel_apply_header_freeze(wb)
    excel_apply_cell_number_formats(wb)
    excel_apply_column_widths(wb)
    excel_apply_conditional_formatting(wb)
    excel_apply_borders(wb)
    excel_apply_hide_columns(wb)
    excel_apply_filter_setup(wb)
    excel_apply_sheet_order(wb)

    # save workbook (.xlsx file)
    wb.remove(ws1)  # remove blank worksheet
    wb.save(filename=dest_filename)


if __name__ == "__main__":
    main()
