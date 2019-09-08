"""Create DFS spreadsheet from stats """

import csv
import json
import re
from os import makedirs, path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, colors
from openpyxl.utils import get_column_letter


def style_range(
    worksheet, cell_range, border=Border(), fill=None, font=None, alignment=None
):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param worksheet:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = worksheet[cell_range.split(":")[0]]
    if alignment:
        worksheet.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = worksheet[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def create_sheet_header(workbook, title, header):
    workbook.create_sheet(title=title)
    workbook[title].append(header)


def pull_data(filename, endpoint):
    """Either pull file from API or from file."""
    data = None
    if not path.isfile(filename):
        print(
            "{} does not exist. Pulling from endpoint [{}]".format(filename, endpoint)
        )
        # send GET request
        response = requests.get(endpoint)
        status = response.status_code

        # if not successful, raise an exception
        if status != 200:
            raise Exception("Requests status != 200. It is: {0}".format(status))

        # store response
        data = response.json()

        # dump json to file for future use to avoid multiple API pulls
        with open(filename, "w") as outfile:
            json.dump(data, outfile)
    else:
        print("File exists [{}]. Nice!".format(filename))
        # load json from file
        with open(filename, "r") as json_file:
            data = json.load(json_file)

    return data


def pull_soup_data(filename, endpoint):
    """Either pull file from html or from file."""
    soup = None
    if not path.isfile(filename):
        print(
            "{} does not exist. Pulling from endpoint [{}]".format(filename, endpoint)
        )
        headers = {
            "Accept": "*/*",
            "Accept-Encoding": "gzip, deflate, sdch",
            "Accept-Language": "en-US,en;q=0.8",
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "Pragma": "no-cache",
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/48.0.2564.97 Safari/537.36"
            ),
        }
        # send GET request
        response = requests.get(endpoint, headers=headers)
        status = response.status_code

        # if not successful, raise an exception
        if status != 200:
            raise Exception("Requests status != 200. It is: {0}".format(status))

        # dump html to file to avoid multiple requests
        with open(filename, "w") as outfile:
            print(response.text, file=outfile)

        soup = BeautifulSoup(response.text, "html5lib")
    else:
        print("File exists [{}]. Nice!".format(filename))
        # load html from file
        with open(filename, "r") as html_file:
            soup = BeautifulSoup(html_file, "html5lib")

    return soup


def get_nfl_snaps(workbook):
    """Retrieve snaps from lineups.com API."""
    endpoint = "https://api.lineups.com/nfl/fetch/snaps/2018/OFF"
    filename = "nfl_snaps.json"
    directory = "sources"
    full_path = path.join(directory, filename)

    # if file doesn't exist, let's pull it. otherwise - use the file.
    data = pull_data(full_path, endpoint)

    if data is None:
        raise Exception("Failed to pull data from API or file.")

    player_data = data["data"]

    # create worksheet
    title = "SNAPS"
    header = [
        "name",
        "position",
        "team",
        "season average",
        "week1",
        "week2",
        "week3",
        "week4",
        "week5",
        "week6",
        "week7",
        "week8",
        "week9",
        "week10",
        "week11",
        "week12",
        "week13",
        "week14",
        "week15",
        "week16",
    ]
    create_sheet_header(workbook, title, header)

    for data in player_data:
        name = data["full_name"]
        position = data["position"]
        team = data["team"]
        weeks = data["snap_percentage_by_week"]  # list
        season_average = data["season_snap_percent"]

        # we only care about RB/TE/WR
        if position not in ["RB", "TE", "WR"]:
            continue

        # remove '.' from name
        name = name.replace(".", "")

        # convert weeks dict to list
        all_weeks = conv_weeks_to_padded_list(weeks)

        # add three lists together
        pre_weeks = [name, position, team, season_average]
        # post_weeks = [targets, average, recv_touchdowns]

        workbook[title].append(pre_weeks + all_weeks)


def get_nfl_targets(workbook):
    """Retrieve targets from lineups.com API."""
    endpoint = "https://api.lineups.com/nfl/fetch/targets/2018/OFF"
    filename = "nfl_targets.json"
    directory = "sources"
    full_path = path.join(directory, filename)

    # if file doesn't exist, let's pull it. otherwise - use the file.
    data = pull_data(full_path, endpoint)

    player_data = data["data"]

    # create worksheet
    title = "TARGETS"
    header = [
        "name",
        "position",
        "team",
        "season average",
        "week1",
        "week2",
        "week3",
        "week4",
        "week5",
        "week6",
        "week7",
        "week8",
        "week9",
        "week10",
        "week11",
        "week12",
        "week13",
        "week14",
        "week15",
        "week16",
        "targets",
        "recv touchdowns",
    ]
    create_sheet_header(workbook, title, header)

    for data in player_data:
        # TODO target percentage? it's by week as well
        name = data["full_name"]
        position = data["position"]
        team = data["team"]
        targets = data["total"]
        weeks = data["weeks"]  # dict
        season_average = data["average"]
        recv_touchdowns = data["receiving_touchdowns"]
        catch_percentage = data["catch_percentage"]
        season_target_percent = data["season_target_percent"]

        # we only care about RB/TE/WR
        if position not in ["RB", "TE", "WR"]:
            continue

        # remove '.' from name
        name = name.replace(".", "")

        # convert weeks dict to list
        all_weeks = conv_weeks_to_padded_list(weeks)

        # add three lists together
        pre_weeks = [name, position, team, season_average]
        post_weeks = [targets, recv_touchdowns]

        # insert all_weeks list into ls
        # ls = [name, position, rating, team, receptions, average, touchdowns]
        # print("trying to insert: ls[2:{}]".format(len(all_weeks)))
        # ls[4:len(all_weeks)-1] = all_weeks
        # print(ls)

        workbook[title].append(pre_weeks + all_weeks + post_weeks)


def get_nfl_receptions(workbook):
    """Retrieve receptions from lineups.com API."""
    endpoint = "https://api.lineups.com/nfl/fetch/receptions/2018/OFF"
    filename = "nfl_receptions.json"
    directory = "sources"
    full_path = path.join(directory, filename)

    # if file doesn't exist, let's pull it. otherwise - use the file.
    data = pull_data(full_path, endpoint)

    # we just want player data
    player_data = data["data"]

    # create worksheet
    title = "RECEPTIONS"
    header = [
        "name",
        "position",
        "team",
        "season average",
        "week1",
        "week2",
        "week3",
        "week4",
        "week5",
        "week6",
        "week7",
        "week8",
        "week9",
        "week10",
        "week11",
        "week12",
        "week13",
        "week14",
        "week15",
        "week16",
        "receptions",
        "touchdowns",
    ]
    create_sheet_header(workbook, title, header)

    for data in player_data:
        name = data["name"]
        position = data["position"]
        team = data["team"]
        receptions = data["receptions"]
        weeks = data["weeks"]  # dict
        season_average = data["average"]
        touchdowns = data["touchdowns"]

        # we only care about RB/TE/WR
        if position not in ["RB", "TE", "WR"]:
            continue

        # remove '.' from name
        name = name.replace(".", "")

        # convert weeks dict to list
        all_weeks = conv_weeks_to_padded_list(weeks)

        # add three lists together
        pre_weeks = [name, position, team, season_average]
        post_weeks = [receptions, touchdowns]

        workbook[title].append(pre_weeks + all_weeks + post_weeks)


def get_nfl_rush_atts(workbook):
    """Retrieve receptions from lineups.com API."""
    endpoint = "https://api.lineups.com/nfl/fetch/rush/2018/OFF"
    filename = "nfl_rush_atts.json"
    directory = "sources"
    full_path = path.join(directory, filename)

    # if file doesn't exist, let's pull it. otherwise - use the file.
    data = pull_data(full_path, endpoint)

    # we just want player data
    player_data = data["data"]

    # create worksheet
    title = "RUSH_ATTS"
    header = [
        "name",
        "position",
        "team",
        "season average",
        "week1",
        "week2",
        "week3",
        "week4",
        "week5",
        "week6",
        "week7",
        "week8",
        "week9",
        "week10",
        "week11",
        "week12",
        "week13",
        "week14",
        "week15",
        "week16",
        "attempts",
        "touchdowns",
    ]
    create_sheet_header(workbook, title, header)

    for data in player_data:
        # TODO rushing_attempt_percentage_by_week
        name = data["name"]
        position = data["position"]
        team = data["team"]
        attempts = data["total"]
        weeks = data["weeks"]  # dict
        season_average = data["average"]
        touchdowns = data["touchdowns"]

        # we only care about QB/RB/WR
        if position not in ["QB", "RB", "WR"]:
            continue

        # remove '.' from name
        name = name.replace(".", "")

        # convert weeks dict to list
        all_weeks = conv_weeks_to_padded_list(weeks)

        # add three lists together
        pre_weeks = [name, position, team, season_average]
        post_weeks = [attempts, touchdowns]
        ls = pre_weeks + all_weeks + post_weeks

        workbook[title].append(ls)


def get_nfl_def_stats(workbook):
    # https://www.lineups.com/nfl/teams/stats/defense-stats
    # get passing yds/att
    # td / att (td %)
    # att / completion (compl %)
    """Retrieve receptions from lineups.com API."""
    endpoint = "https://api.lineups.com/nfl/fetch/teams/stats/defense-stats/current"
    filename = "nfl_def_stats.json"
    directory = "sources"
    full_path = path.join(directory, filename)

    # if file doesn't exist, let's pull it. otherwise - use the file.
    data = pull_data(full_path, endpoint)

    # we just want player data
    player_data = data["data"]

    # create worksheet
    title = "DEF_STATS"
    header = [
        "team abbv",
        "team",
        "pass_att",
        "pass_yd_per_att",
        "pass_compls",
        "pass_yd_per_compl",
        "pass_yds",
        "pass_tds",
        "pass_td_per_att",
        "compl_perc",
    ]
    create_sheet_header(workbook, title, header)

    team_map = {
        "Atlanta Falcons": "ATL",
        "Indianapolis Colts": "IND",
        "San Francisco 49ers": "SF",
        "Oakland Raiders": "OAK",
        "Tampa Bay Buccaneers": "TB",
        "Kansas City Chiefs": "KC",
        "New York Giants": "NYG",
        "Cincinnati Bengals": "CIN",
        "Pittsburgh Steelers": "PIT",
        "Denver Broncos": "DEN",
        "Cleveland Browns": "CLE",
        "New England Patriots": "NE",
        "Minnesota Vikings": "MIN",
        "Miami Dolphins": "MIA",
        "Green Bay Packers": "GB",
        "Los Angeles Chargers": "LAC",
        "New Orleans Saints": "NO",
        "New York Jets": "NYJ",
        "Arizona Cardinals": "ARI",
        "Buffalo Bills": "BUF",
        "Houston Texans": "HOU",
        "Detroit Lions": "DET",
        "Jacksonville Jaguars": "JAX",
        "Los Angeles Rams": "LAR",
        "Seattle Seahawks": "SEA",
        "Philadelphia Eagles": "PHI",
        "Carolina Panthers": "CAR",
        "Tennessee Titans": "TEN",
        "Washington Redskins": "WAS",
        "Dallas Cowboys": "DAL",
        "Chicago Bears": "CHI",
        "Baltimore Ravens": "BAL",
    }

    for data in player_data:
        # TODO rushing_attempt_percentage_by_week
        team = data["team"]
        team_abbv = team_map[team]
        pass_att = data["passing_attempts"]
        pass_yd_per_att = data["passing_yards_per_attempt"]
        pass_compls = data["passing_completions"]
        pass_yd_per_compl = data["passing_yards_per_completion"]
        pass_yds = data["passing_yards"]
        pass_tds = data["passing_touchdowns"]

        # personal
        try:
            pass_td_per_att = "{0:.4f}".format(pass_tds / pass_att)
        except ZeroDivisionError:
            pass_td_per_att = "{0:.4f}".format(0)

        try:
            compl_perc = "{0:.4f}".format(pass_compls / pass_att)
        except ZeroDivisionError:
            compl_perc = "{0:.4f}".format(0)

        # remove '.' from name
        # name = name.replace('.', '')

        workbook[title].append(
            [
                team_abbv,
                team,
                pass_att,
                pass_yd_per_att,
                pass_compls,
                pass_yd_per_compl,
                pass_yds,
                pass_tds,
                compl_perc,
                pass_td_per_att,
            ]
        )


def conv_weeks_to_padded_list(weeks):
    """Convert weeks dict or list to padded list (16 weeks)."""
    all_weeks = []
    if isinstance(weeks, list):
        for week in weeks:
            if week is None:
                all_weeks.append("")
            else:
                all_weeks.append(week)
    elif isinstance(weeks, dict):
        for i in range(0, len(weeks)):
            # if weeks is None, put in blank string
            # 0 would mean they played but didn't get a snap
            if weeks[str(i + 1)] is None:
                all_weeks.append("")
            else:
                all_weeks.append(weeks[str(i + 1)])

    # pad weeks to 16 (a = [])
    # more visual/pythonic
    # a = (a + N * [''])[:N]
    var = 16
    all_weeks = (all_weeks + var * [""])[:var]

    return all_weeks


# def get_vegas_ows(workbook):
#     endpoint = "https://rotogrinders.com/schedules/nfl"

#     filename = "vegas_script.html"
#     directory = "sources"
#     full_path = path.join(directory, filename)

#     # create worksheet
#     title = "VEGAS"
#     header = [
#         "Time",
#         "Team",
#         "Opponent",
#         "Line",
#         "MoneyLine",
#         "Over/Under",
#         "Projected Points",
#         "Projected Points Change",
#     ]
#     create_sheet_header(workbook, title, header)

#     # pull data
#     soup = pull_soup_data(full_path, endpoint)

#     # find script(s) in the html
#     script = soup.findAll("script")

#     js_vegas_data = script[11].string

#     # replace dumb names
#     js_vegas_data = js_vegas_data.replace("KCC", "KC")
#     js_vegas_data = js_vegas_data.replace("JAC", "JAX")

#     pattern = re.compile(r"data = (.*);")

#     json_str = pattern.search(js_vegas_data).group(1)
#     vegas_json = json.loads(json_str)
#     for matchup in vegas_json:
#         workbook[title].append(
#             [
#                 matchup["time"]["display"],
#                 matchup["team"],
#                 matchup["opponent"],
#                 matchup["line"],
#                 matchup["moneyline"],
#                 matchup["overunder"],
#                 matchup["projected"],
#                 matchup["projectedchange"]["value"],
#             ]
#         )


def get_vegas_ows(workbook):
    endpoint = "https://www.oneweekseason.com/"

    filename = "vegas_script.html"
    directory = "sources"
    full_path = path.join(directory, filename)

    # create worksheet
    title = "VEGAS"
    header = [
        "Time",
        "Team",
        "Opponent",
        "Line",
        "MoneyLine",
        "Over/Under",
        "Projected Points",
        "Projected Points Change",
    ]
    create_sheet_header(workbook, title, header)

    # pull data
    soup = pull_soup_data(full_path, endpoint)

    # find script(s) in the html
    # script = soup.findAll("script")
    # sidebar_wrapper = soup.findAll("div", {"class": "sidebar-matchup-wrapper"})
    matchup_divs = soup.findAll("div", {"class": "game-matchup"})

    matchups = []

    for div in matchup_divs:
        home_team, home_total = div.find(class_="left-team").text.split()
        away_team, away_total = div.find(class_="right-team").text.split()
        total = home_total + away_total

        matchups.append(
            {
                "home_team": home_team,
                "home_total": home_total,
                "away_team": away_team,
                "away_total": away_total,
                "total": total,
            }
        )

    for matchup in matchups:
        workbook[title].append(
            [
                matchup["home_team"],
                matchup["home_total"],
                matchup["away_team"],
                matchup["away_total"],
                matchup["total"],
            ]
        )


def get_dvoa_rankings(workbook):
    endpoint = "https://www.footballoutsiders.com/stats/teamdef"
    filename = "html_defense.html"
    directory = "sources"
    full_path = path.join(directory, filename)

    # pull data
    soup = pull_soup_data(full_path, endpoint)

    # find all tables (3) in the html
    table = soup.findAll("table")

    if table:
        # create worksheet
        title = "TEAMDEF"
        workbook.create_sheet(title=title)

        defense_stats = table[0]

        # find header
        table_header = defense_stats.find("thead")
        # there is one header row
        header_row = table_header.find("tr")
        # loop through header columns and append to worksheet
        header_cols = header_row.find_all("th")
        header = [ele.text.strip() for ele in header_cols]
        workbook[title].append(header)

        # find the rest of the table header_rows
        rows = defense_stats.find_all("tr")
        for row in rows:
            cols = row.find_all("td")
            cols = [ele.text.strip() for ele in cols]
            if cols:
                workbook[title].append(cols)

        # separate function for second table
        get_dvoa_recv_rankings(workbook, table[1], title)


def get_dvoa_recv_rankings(workbook, soup_table, title):
    # VS types of receivers
    def_recv_stats = soup_table
    table_header = def_recv_stats.find("thead")
    header_rows = table_header.find_all("tr")

    # style for merge + center
    alignment = Alignment(horizontal="center", vertical="center")

    # there are two header rows
    for i, row in enumerate(header_rows):
        header_cols = row.find_all("th")
        header = [ele.text.strip() for ele in header_cols]
        # first header row has some merged cells
        if i == 0:
            # merge + center
            workbook[title]["C35"] = header[2]  # vs. WR1
            workbook[title].merge_cells("C35:F35")
            style_range(workbook[title], "C35:F35", alignment=alignment)
            workbook[title]["G35"] = header[3]  # vs. WR2
            workbook[title].merge_cells("G35:J35")
            style_range(workbook[title], "G35:J35", alignment=alignment)
            workbook[title]["K35"] = header[4]  # vs. OTHER
            workbook[title].merge_cells("K35:N35")
            style_range(workbook[title], "K35:N35", alignment=alignment)
            workbook[title]["O35"] = header[5]  # vs. TE
            workbook[title].merge_cells("O35:R35")
            style_range(workbook[title], "O35:R35", alignment=alignment)
            workbook[title]["S35"] = header[6]  # vs. RB
            workbook[title].merge_cells("S35:V35")
            style_range(workbook[title], "S35:V35", alignment=alignment)
        elif i == 1:
            workbook[title].append(header)
        # for c in cols:
        #     print(c.get_text(strip=True))
        # print(cols)

        # create_sheet_header(workbook, title, header)
        # print(header)

    rows = def_recv_stats.find_all("tr")
    for row in rows:
        cols = row.find_all("td")
        cols = [ele.text.strip() for ele in cols]
        if cols:
            workbook[title].append(cols)


def get_oline_rankings(workbook):
    endpoint = "https://www.footballoutsiders.com/stats/ol"
    file_name = "html_oline.html"
    directory = "sources"
    full_path = path.join(directory, file_name)

    # pull data
    soup = pull_soup_data(full_path, endpoint)

    # find all tables (2) in the html
    table = soup.findAll("table")

    if table:
        # create worksheet
        title = "OLINE"
        workbook.create_sheet(title=title)

        oline_stats = table[0]

        # find header
        table_header = oline_stats.find("thead")
        # there is one header row
        header_row = table_header.find("tr")
        # loop through header columns and append to worksheet
        header_cols = header_row.find_all("th")
        header = [ele.text.strip() for ele in header_cols]
        workbook[title].append(header)

        # find the rest of the table header_rows
        rows = oline_stats.find_all("tr")
        for row in rows:
            cols = row.find_all("td")
            cols = [ele.text.strip() for ele in cols]
            if cols:
                workbook[title].append(cols)


def get_dline_rankings(workbook):
    endpoint = "https://www.footballoutsiders.com/stats/dl"
    filename = "html_dline.html"
    directory = "sources"
    full_path = path.join(directory, filename)

    # pull data
    soup = pull_soup_data(full_path, endpoint)

    # find all tables (2) in the html
    table = soup.findAll("table")

    if table:
        # create worksheet
        title = "DLINE"
        workbook.create_sheet(title=title)

        oline_stats = table[0]

        # find header
        table_header = oline_stats.find("thead")
        # there is one header row
        header_row = table_header.find("tr")
        # loop through header columns and append to worksheet
        header_cols = header_row.find_all("th")
        header = [ele.text.strip() for ele in header_cols]
        workbook[title].append(header)

        # find the rest of the table header_rows
        rows = oline_stats.find_all("tr")
        for row in rows:
            cols = row.find_all("td")
            cols = [ele.text.strip() for ele in cols]
            if cols:
                workbook[title].append(cols)


def get_qb_stats_outsiders(workbook):
    endpoint = "https://www.footballoutsiders.com/stats/qb"
    filename = "html_qb.html"
    directory = "sources"
    full_path = path.join(directory, filename)

    # pull data
    soup = pull_soup_data(full_path, endpoint)

    # find all tables (3) in the html
    tables = soup.findAll("table")

    if tables:
        # create worksheet
        title = "QB_STATS"
        workbook.create_sheet(title=title)

        for table in tables:
            qb_stats = table

            # find header
            table_header = qb_stats.find("thead")
            # there is one header row
            header_row = table_header.find("tr")
            # loop through header columns and append to worksheet
            header_cols = header_row.find_all("th")
            header = [ele.text.strip() for ele in header_cols]
            workbook[title].append(header)

            # find the rest of the table header_rows
            rows = qb_stats.find_all("tr")
            for row in rows:
                cols = row.find_all("td")
                cols = [ele.text.strip() for ele in cols]
                if cols:
                    workbook[title].append(cols)


def fpros_ecr(workbook, position):
    if position == "QB" or position == "DST":
        endpoint = "https://www.fantasypros.com/nfl/rankings/{}.php".format(
            position.lower()
        )
    else:
        endpoint = "https://www.fantasypros.com/nfl/rankings/ppr-{}.php".format(
            position.lower()
        )

    filename = "ecr_{}.html".format(position)
    directory = "sources"
    full_path = path.join(directory, filename)

    # pull data
    soup = pull_soup_data(full_path, endpoint)

    # find all tables (2) in the html
    table = soup.find("table", id="rank-data")

    if table:
        # create worksheet
        title = "{0}_ECR".format(position)
        workbook.create_sheet(title=title)

        # # find header
        table_header = table.find("thead")
        # there is one header row
        header_row = table_header.find("tr")
        # loop through header columns and append to worksheet
        header_cols = header_row.find_all("th")
        header = [ele.text.strip() for ele in header_cols]
        workbook[title].append(header)

        # find the rest of the table header_rows
        rows = table.find_all("tr")
        for row in rows:
            cols = row.find_all("td")
            # cols = [ele.text.strip() for ele in cols]
            # change from list comp for just fpros
            new_cols = []
            for i, ele in enumerate(cols):

                if i == 2:  # name of player
                    txt = ele.find(class_="full-name").text

                    # remove periods (T.J. Yeldon, T.Y. Hilton)
                    txt = txt.replace(".", "")
                else:
                    txt = ele.text.strip()
                # replace JAX
                txt = txt.replace("JAC", "JAX")

                # really? just to fix mitchell?
                if position == "QB":
                    txt = txt.replace("Mitch", "Mitchell")
                new_cols.append(txt)
            if new_cols:
                workbook[title].append(new_cols)


def position_tab(workbook, values, title, fdraft_dict=None):
    # create positional tab if it does not exist
    # and set header(s)
    if title not in workbook.sheetnames:
        workbook.create_sheet(title=title)

        # style for merge + center
        alignment = Alignment(horizontal="center", vertical="center")

        # second header
        all_positions_header = [
            "Position",
            "Name",
            "Opp",
            "Abbv",
            "Salary",
            "Salary%",
            "Implied Total",
            "O/U",
            "Line",
        ]

        # set row height
        workbook[title].row_dimensions[2].height = 40

        # more header fields based on position
        position_fields = []
        if title == "QB":
            top_lvl_header(workbook, title, "DK", "E", 1, "FF000000")
            top_lvl_header(workbook, title, "VEGAS", "G", 2, "FFFFC000")
            top_lvl_header(workbook, title, "SEASON", "J", 2, "FF5B9BD5")
            top_lvl_header(workbook, title, "PRESSURE", "M", 1, "FF00B0F0")
            top_lvl_header(workbook, title, "MATCHUP", "O", 2, "FFED7D31")
            top_lvl_header(workbook, title, "RANKINGS", "R", 2, "FF70AD47")
            top_lvl_header(workbook, title, "FDRAFT", "W", 1, "FFA8F3D9")

            position_fields = [
                "Rushing Yards",
                "DYAR",
                "QBR",
                "O-Line Sack%",
                "D-Line Sack%",
                "Def Yds/Att",
                "Def Comp%",
                "Def TD%",
                "Ave PPG",
                "ECR",
                "+/- Rank",
                "ECR Data",
                "Salary Rank",
                "FD Salary",
                "FD Salary%",
            ]
        elif title == "RB":
            top_lvl_header(workbook, title, "DK", "E", 1, "FF000000")
            top_lvl_header(workbook, title, "VEGAS", "G", 2, "FFFFC000")
            top_lvl_header(workbook, title, "MATCHUP", "J", 3, "FFED7D31")
            top_lvl_header(workbook, title, "SEASON", "N", 2, "FF5B9BD5")
            top_lvl_header(workbook, title, "LAST WEEK", "Q", 2, "FF4472C4")
            top_lvl_header(workbook, title, "RANKINGS", "T", 2, "FF70AD47")
            top_lvl_header(workbook, title, "FDRAFT", "Y", 1, "FFA8F3D9")

            position_fields = [
                "Run DVOA",
                "Pass DVOA",
                "O-Line",
                "D-Line",
                "Snap%",
                "Rush ATTs",
                "Targets",
                "Snap%",
                "Rush ATTs",
                "Targets",
                "Ave PPG",
                "ECR",
                "+/- Rank",
                "ECR Data",
                "Salary Rank",
                "FD Salary",
                "FD Salary%",
            ]
        elif title == "WR":
            top_lvl_header(workbook, title, "DK", "E", 1, "FF000000")
            top_lvl_header(workbook, title, "VEGAS", "G", 2, "FFFFC000")
            top_lvl_header(workbook, title, "MATCHUP", "J", 2, "FFED7D31")
            top_lvl_header(workbook, title, "SEASON", "M", 2, "FF5B9BD5")
            top_lvl_header(workbook, title, "LAST WEEK", "P", 2, "FF4472C4")
            top_lvl_header(workbook, title, "RANKINGS", "S", 2, "FF70AD47")
            top_lvl_header(workbook, title, "FDRAFT", "X", 1, "FFA8F3D9")

            position_fields = [
                "Pass DVOA",
                "vs. WR1",
                "vs. WR2",
                "Snap%",
                "Targets",
                "Recepts",
                "Snap%",
                "Targets",
                "Recepts",
                "Ave PPG",
                "ECR",
                "+/- Rank",
                "ECR Data",
                "Salary Rank",
                "FD Salary",
                "FD Salary%",
            ]
        elif title == "TE":
            top_lvl_header(workbook, title, "DK", "E", 1, "FF000000")
            top_lvl_header(workbook, title, "VEGAS", "G", 2, "FFFFC000")
            top_lvl_header(workbook, title, "MATCHUP", "J", 1, "FFED7D31")
            top_lvl_header(workbook, title, "SEASON", "L", 1, "FF5B9BD5")
            top_lvl_header(workbook, title, "LAST WEEK", "N", 1, "FF4472C4")
            top_lvl_header(workbook, title, "RANKINGS", "P", 2, "FF70AD47")
            top_lvl_header(workbook, title, "FDRAFT", "U", 1, "FFA8F3D9")

            position_fields = [
                "Pass DVOA",
                "vs. TE",
                "Snap%",
                "Targets",
                "Snap%",
                "Targets",
                "Ave PPG",
                "ECR",
                "+/- Rank",
                "ECR Data",
                "Salary Rank",
                "FD Salary",
                "FD Salary%",
            ]
        elif title == "DST":
            top_lvl_header(workbook, title, "DK", "E", 1, "FF000000")
            top_lvl_header(workbook, title, "VEGAS", "G", 2, "FFFFC000")
            top_lvl_header(workbook, title, "RANKINGS", "J", 2, "FF70AD47")

            position_fields = [
                "Ave PPG",
                "ECR",
                "+/- Rank",
                "ECR Data",
                "Salary Rank",
                "FD Salary",
                "FD Salary%",
            ]

        # find max row to append
        append_row = workbook[title].max_row + 1
        header = all_positions_header + position_fields

        # change row font and alignment
        font = Font(b=True, color="FF000000")
        alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        # just set for row range
        rng = "{0}:{1}".format(2, 2)
        for cell in workbook[title][rng]:
            cell.font = font
            cell.alignment = alignment

        for i, field in enumerate(header):
            workbook[title].cell(row=append_row, column=i + 1, value=field)

    keys = [
        "pos",
        "name_id",
        "name",
        "id",
        "roster_pos",
        "salary",
        "matchup",
        "abbv",
        "avg_ppg",
    ]
    stats_dict = dict(zip(keys, values))
    stats_dict["salary_perc"] = "{0:.1%}".format(float(stats_dict["salary"]) / 50000)

    # 'fix' name to remove extra stuff like Jr or III (Todd Gurley II for example)
    name = " ".join(stats_dict["name"].split(" ")[:2])
    # also remove periods (T.J. Yeldon for example)
    name = name.replace(".", "")
    stats_dict["name"] = name

    # find opp, opp_excel, and game_time
    home_at_away, game_time = stats_dict["matchup"].split(" ", 1)
    stats_dict["game_time"] = game_time
    home_team, away_team = home_at_away.split("@")
    if stats_dict["abbv"] == home_team:
        stats_dict["opp"] = away_team
        stats_dict["opp_excel"] = "vs. {}".format(away_team)
    else:
        stats_dict["opp"] = home_team
        stats_dict["opp_excel"] = "at {}".format(home_team)

    # find max row to append
    append_row = workbook[title].max_row + 1

    # insert rows of data
    all_positions_fields = [
        stats_dict["pos"],
        stats_dict["name"],
        stats_dict["opp_excel"],
        stats_dict["abbv"],
        stats_dict["salary"],
        stats_dict["salary_perc"],
        bld_excel_formula(
            "VEGAS", "$G$2:$G$29", "$D", append_row, "$B$2:$B$29"
        ),  # implied total
        bld_excel_formula(
            "VEGAS", "$F$2:$F$29", "$D", append_row, "$B$2:$B$29"
        ),  # over/under
        bld_excel_formula(
            "VEGAS", "$D$2:$D$29", "$D", append_row, "$B$2:$B$29"
        ),  # line
    ]

    # more header fields based on position
    positional_fields = []

    # get max_row from position ECR tab
    max_row = workbook[title + "_ECR"].max_row + 1
    if title == "QB":
        positional_fields = [
            # rushing yards
            bld_excel_formula(
                "QB_STATS",
                "$K$44:$K$82",
                "$B",
                append_row,
                "$A$44:$A$82",
                qb_stats=True,
            ),
            # DYAR
            bld_excel_formula(
                "QB_STATS", "$C$2:$C$42", "$B", append_row, "$A$2:$A$42", qb_stats=True
            ),
            # QBR
            bld_excel_formula(
                "QB_STATS", "$J$2:$J$35", "$B", append_row, "$A$2:$A$35", qb_stats=True
            ),
            # o-line
            bld_excel_formula("OLINE", "P$2:$P$35", "$D", append_row, "$B$2:$B$33"),
            # d-line
            bld_excel_formula(
                "DLINE", "P$2:$P$35", "$C", append_row, "$B$2:$B$33", right=True
            ),
            # matchup passing_yards_per_attempt
            bld_excel_formula(
                "DEF_STATS",
                "D$2:$D${0}".format(max_row),
                "$C",
                append_row,
                "$A$2:$A${0}".format(max_row),
                right=True,
            ),
            # matchup compl %
            bld_excel_formula(
                "DEF_STATS",
                "I$2:$I${0}".format(max_row),
                "$C",
                append_row,
                "$A$2:$A${0}".format(max_row),
                right=True,
            ),
            # matchup td %
            bld_excel_formula(
                "DEF_STATS",
                "J$2:$J${0}".format(max_row),
                "$C",
                append_row,
                "$A$2:$A${0}".format(max_row),
                right=True,
            ),
            # Ave PPG
            stats_dict["avg_ppg"],
            # ECR
            "=RANK(U{0}, $U$3:$U${1},1)".format(append_row, max_row),
            # + / - salary
            "=V{0} - S{0}".format(append_row),
            # ECR DATA
            bld_excel_formula(
                "QB_ECR",
                "$A$2:$A${}".format(max_row),
                "$B",
                append_row,
                "$C$2:$C${}".format(max_row),
            ),
            # salary rank (low to high)
            "=RANK(E{0}, $E$3:$E${1},0)".format(append_row, max_row),
        ]
        # fdraft salary
        if fdraft_dict and name in fdraft_dict:
            positional_fields.append(fdraft_dict[name]["salary"])
            # fdraft salary perc
            positional_fields.append(fdraft_dict[name]["salary_perc"])
        # style column L & M (pressure %) with %/decimals
        for cell in workbook[title]["M"]:
            cell.number_format = "##0.0%"
        for cell in workbook[title]["N"]:
            cell.number_format = "##0.0%"
        # style column L & M (matchup %) with %/decimals
        for cell in workbook[title]["P"]:
            cell.number_format = "##0.0%"
        for cell in workbook[title]["Q"]:
            cell.number_format = "##0.0%"

    elif title == "RB":
        max_row = workbook[title + "_ECR"].max_row
        positional_fields = [
            # run dvoa
            # 'x',  # for testing bld_excel_formula_2
            bld_excel_formula(
                "TEAMDEF", "J$2:$J$33", "$C", append_row, "$B$2:$B$33", right=True
            ),
            # pass dvoa (vs. RB)
            bld_excel_formula(
                "TEAMDEF", "T$37:$T$68", "$C", append_row, "$B$37:$B$68", right=True
            ),
            # o line
            bld_excel_formula("OLINE", "$C$2:$C$33", "$D", append_row, "$B$2:$B$33"),
            # d line
            bld_excel_formula(
                "DLINE", "$C$2:$C$33", "$C", append_row, "$B$2:$B$33", right=True
            ),
            # season snap%
            bld_excel_formula("SNAPS", "$D$2:$D$449", "$B", append_row, "$A$2:$A$449"),
            # season rush atts
            bld_excel_formula(
                "RUSH_ATTS", "$D$2:$D$449", "$B", append_row, "$A$2:$A$449"
            ),
            # season targets
            bld_excel_formula(
                "TARGETS", "$D$2:$D$449", "$B", append_row, "$A$2:$A$449"
            ),
            # week snap% (week6)
            bld_excel_formula(
                "SNAPS", "$K$2:$K$449", "$B", append_row, "$A$2:$A$449", week=True
            ),
            # week rush atts (week 6)
            bld_excel_formula(
                "RUSH_ATTS", "$K$2:$K$449", "$B", append_row, "$A$2:$A$449", week=True
            ),
            # week targets (week 6)
            bld_excel_formula(
                "TARGETS", "$K$2:$K$449", "$B", append_row, "$A$2:$A$449", week=True
            ),
            # Ave PPG
            stats_dict["avg_ppg"],
            # ECR
            "=RANK(V{0}, $V$3:$V$69,1)".format(append_row),
            # +/- rank
            "x",
            # ECR Data
            bld_excel_formula(
                "RB_ECR",
                "$A$2:$A${}".format(max_row),
                "$B",
                append_row,
                "$C$2:$C${}".format(max_row),
            ),
            # salary rank
            "x",
        ]
    elif title == "WR":
        positional_fields = [
            # pass dvoa
            bld_excel_formula(
                "TEAMDEF", "$H$2:$H$34", "$C", append_row, "$B$2:$B$34", right=True
            ),
            # vs. WR1
            bld_excel_formula(
                "TEAMDEF", "$D$37:$D$68", "$C", append_row, "$B$37:$B$68", right=True
            ),
            # vs. WR2
            bld_excel_formula(
                "TEAMDEF", "$H$37:$H$68", "$C", append_row, "$B$37:$B$68", right=True
            ),
            # season snap%
            bld_excel_formula("SNAPS", "$D$2:$D$449", "$B", append_row, "$A$2:$A$449"),
            # season targets
            bld_excel_formula(
                "TARGETS", "$D$2:$D$449", "$B", append_row, "$A$2:$A$449"
            ),
            # season receptions
            bld_excel_formula(
                "RECEPTIONS", "$D$2:$D$449", "$B", append_row, "$A$2:$A$449"
            ),
            # week snap% (week6)
            bld_excel_formula(
                "SNAPS", "$K$2:$K$449", "$B", append_row, "$A$2:$A$449", week=True
            ),
            # week targets (week 6)
            bld_excel_formula(
                "TARGETS", "$K$2:$K$449", "$B", append_row, "$A$2:$A$449", week=True
            ),
            # week targets (week 6)
            bld_excel_formula(
                "RECEPTIONS", "$K$2:$K$449", "$B", append_row, "$A$2:$A$449", week=True
            ),
            # Ave PPG
            stats_dict["avg_ppg"],
            # ECR
            "=RANK(S{0}, $S$3:$S$94,1)".format(append_row),
            # +/- rank
            "x",
            # ECR Data
            bld_excel_formula(
                "WR_ECR",
                "$A$2:$A${}".format(max_row),
                "$B",
                append_row,
                "$C$2:$C${}".format(max_row),
            ),
            # salary rank
            "x",
        ]
    elif title == "TE":
        positional_fields = [
            # pass dvoa
            bld_excel_formula(
                "TEAMDEF", "$H$2:$H$34", "$C", append_row, "$B$2:$B$34", right=True
            ),
            # vs. TE
            bld_excel_formula(
                "TEAMDEF", "$P$37:$P$68", "$C", append_row, "$B$37:$B$68", right=True
            ),
            # season snap%
            bld_excel_formula("SNAPS", "$D$2:$D$449", "$B", append_row, "$A$2:$A$449"),
            # season targets
            bld_excel_formula(
                "TARGETS", "$D$2:$D$449", "$B", append_row, "$A$2:$A$449"
            ),
            # week snap% (week6)
            bld_excel_formula(
                "SNAPS", "$K$2:$K$449", "$B", append_row, "$A$2:$A$449", week=True
            ),
            # week targets (week 6)
            bld_excel_formula(
                "TARGETS", "$K$2:$K$449", "$B", append_row, "$A$2:$A$449", week=True
            ),
            # Ave PPG
            stats_dict["avg_ppg"],
            # ECR
            "=RANK(R{0}, $R$3:$R$52,1)".format(append_row),
            # +/- rank
            "x",
            # ECR Data
            bld_excel_formula(
                "TE_ECR",
                "$A$2:$A${}".format(max_row),
                "$B",
                append_row,
                "$C$2:$C${}".format(max_row),
            ),
            # salary rank
            "x",
        ]
    elif title == "DST":
        positional_fields = [
            # Ave PPG
            stats_dict["avg_ppg"],
            # ECR
            "=RANK(L{0}, $L$3:$L$52,1)".format(append_row),
            # +/- rank
            "x",
            # ECR Data
            bld_excel_formula(
                "DST_ECR",
                "$A$2:$A${}".format(max_row),
                "$D",
                append_row,
                "$C$2:$C{}".format(max_row),
                dst=True,
            ),
            # # fdraft salary
            # fdraft_dict[name]['salary'],
            # # fdraft salary perc
            # fdraft_dict[name]['salary_perc'],
        ]

    if fdraft_dict and name in fdraft_dict:
        positional_fields.extend(
            [
                # fdraft salary
                fdraft_dict[name]["salary"],
                # fdraft salary perc
                fdraft_dict[name]["salary_perc"],
            ]
        )

    row = all_positions_fields + positional_fields
    
    # center all cells horzitontally/vertically in row
    for i, text in enumerate(row, start=1):
        nice = workbook[title].cell(row=append_row, column=i, value=text)
        alignment = Alignment(horizontal="center", vertical="center")
        nice.alignment = alignment

    # style column D (salary) with currency
    for cell in workbook[title][find_header_col(workbook[title], "Salary")]:
        cell.number_format = "$#,##0_);($#,##0)"

    # style column E (salary %) with %/decimals
    for cell in workbook[title][find_header_col(workbook[title], "Salary%")]:
        cell.number_format = "##0.0%"

    for cell in workbook[title][find_header_col(workbook[title], "FD Salary")]:
        cell.number_format = "$#,##0_);($#,##0)"

    for cell in workbook[title][find_header_col(workbook[title], "FD Salary%")]:
        cell.number_format = "##0.0%"

    # hide column F (abbv)
    workbook[title].column_dimensions["D"].hidden = True


def find_header_col(worksheet, header_value):
    header_row = 2
    # search through header_row for value
    for cell in worksheet[header_row]:
        if cell.value == header_value:
            return cell.column_letter
    return None


def write_RB_cols(workbook):
    worksheet = workbook["RB"]
    position_fields = [
        "Run DVOA",
        "Pass DVOA",
        "O-Line",
        "D-Line",
        "Snap%",
        "Rush ATTs",
        "Targets",
        "Snap%",
        "Rush ATTs",
        "Targets",
        "Ave PPG",
        "ECR",
        "+/- Rank",
        "ECR Data",
        "Salary Rank",
    ]
    # set max_row for formulas
    max_row = worksheet.max_row

    for field in position_fields:
        header_col = find_header_col(worksheet, field)
        print("field {} is in header column {}".format(field, header_col))
        # run dvoa
        if field == "Run DVOA":
            for cell in worksheet[header_col]:
                # skip header rows
                if cell.row <= 2:
                    continue
                cell.value = bld_excel_formula_2(
                    "TEAMDEF",
                    "$J",
                    2,
                    33,  # return $J in TEAMDEF
                    "$C",
                    cell.row,
                    "$B",
                    2,
                    33,
                    right=True,
                )
        elif field == "Pass DVOA":
            for cell in worksheet[header_col]:
                # skip header rows
                if cell.row <= 2:
                    continue
                cell.value = bld_excel_formula_2(
                    "TEAMDEF",
                    "$T",
                    37,
                    max_row,  # return $J in TEAMDEF
                    "$C",
                    cell.row,
                    "$B",
                    37,
                    68,
                    right=True,
                )
        # # pass dvoa (vs. RB)
        # bld_excel_formula('TEAMDEF', 'T$37:$T$68', '$C', append_row, '$B$37:$B$68', right=True),
        # # o line
        # bld_excel_formula('OLINE', '$C$2:$C$33', '$D', append_row, '$B$2:$B$33'),
        # # d line
        # bld_excel_formula('DLINE', '$C$2:$C$33', '$C', append_row, '$B$2:$B$33', right=True),
        # # season snap%
        # bld_excel_formula('SNAPS', '$D$2:$D$449', '$B', append_row, '$A$2:$A$449'),
        # # season rush atts
        # bld_excel_formula('RUSH_ATTS', '$D$2:$D$449', '$B', append_row, '$A$2:$A$449'),
        # # season targets
        # bld_excel_formula('TARGETS', '$D$2:$D$449', '$B', append_row, '$A$2:$A$449'),
        # # week snap% (week6)
        # bld_excel_formula('SNAPS', '$K$2:$K$449', '$B', append_row, '$A$2:$A$449', week=True),
        # # week rush atts (week 6)
        # bld_excel_formula('RUSH_ATTS', '$K$2:$K$449', '$B', append_row, '$A$2:$A$449', week=True),
        # # week targets (week 6)
        # bld_excel_formula('TARGETS', '$K$2:$K$449', '$B', append_row, '$A$2:$A$449', week=True),
        # # Ave PPG
        # stats_dict['avg_ppg'],
        # # ECR
        # '=RANK(V{0}, $V$3:$V$69,1)'.format(append_row),
        # # +/- rank
        # 'x',
        # # ECR Data
        # bld_excel_formula('RB_ECR', '$A$2:$A${}'.format(max_row), '$B', append_row, '$C$2:$C${}'.format(max_row)),


def top_lvl_header(workbook, title, text, start_col, length, color):
    # style for merge + center
    al = Alignment(horizontal="center", vertical="center")
    # bold font
    font = Font(b=True, color="FFFFFFFF")

    # set cell to start merge + insert text
    cell = "{0}1".format(start_col)
    workbook[title][cell] = text
    # set range to format merged cells
    fmt_range = "{0}1:{1}1".format(start_col, chr(ord(start_col) + length))
    style_range(
        workbook[title],
        fmt_range,
        font=font,
        fill=PatternFill(patternType="solid", fgColor=color),
        alignment=al,
    )


def bld_excel_formula_2(
    title,
    rtrn_col,
    rtrn_start,
    rtrn_stop,
    match,
    row,
    match_col,
    match_start,
    match_stop,
    week=False,
    right=False,
    qb_stats=False,
    dst=False,
):
    # '=INDEX(OLINE!$C$2:$C$33,MATCH($F{0},OLINE!$B$2:$B$33,0))'.format(append_row),

    rtrn_range = "{0}{1}:{0}{2}".format(rtrn_col, rtrn_start, rtrn_stop)
    match_range = "{0}{1}:{0}{2}".format(match_col, match_start, match_stop)
    # use RIGHT for splitting the opponent. IE JAX for vs. JAX
    if right:
        base_formula = 'INDEX({0}!{1}, MATCH(RIGHT({2}{3}, LEN({2}{3}) - SEARCH(" ",{2}{3},1)) & "*", {0}!{4},0))'.format(
            title, rtrn_range, match, row, match_range
        )
    elif qb_stats:
        base_formula = 'INDEX({0}!{1}, MATCH(LEFT({2}{3}, 1) & "*" & RIGHT({2}{3}, LEN({2}{3}) - SEARCH(" ",{2}{3},1)) & "*", {0}!{4},0))'.format(
            title, rtrn_range, match, row, match_range
        )
    elif dst:
        base_formula = 'INDEX({0}!{1}, MATCH("*(" & {2}{3} & "*", {0}!{4},0))'.format(
            title, rtrn_range, match, row, match_range
        )
    else:
        base_formula = 'INDEX({0}!{1}, MATCH({2}{3} & "*", {0}!{4},0))'.format(
            title, rtrn_range, match, row, match_range
        )

    # if we are looking at weekly stats, blank should be blank, not zero
    if week:
        formula = 'IF(ISBLANK({0}), " ", {0})'.format(base_formula)
    else:
        formula = base_formula

    return "=" + formula


def bld_excel_formula(
    title,
    rtrn_range,
    match,
    row,
    match_range,
    week=False,
    right=False,
    qb_stats=False,
    dst=False,
):
    # '=INDEX(OLINE!$C$2:$C$33,MATCH($F{0},OLINE!$B$2:$B$33,0))'.format(append_row),

    # use RIGHT for splitting the opponent. IE JAX for vs. JAX
    if right:
        base_formula = 'INDEX({0}!{1}, MATCH(RIGHT({2}{3}, LEN({2}{3}) - SEARCH(" ",{2}{3},1)) & "*", {0}!{4},0))'.format(
            title, rtrn_range, match, row, match_range
        )
    elif qb_stats:
        base_formula = 'INDEX({0}!{1}, MATCH(LEFT({2}{3}, 1) & "*" & RIGHT({2}{3}, LEN({2}{3}) - SEARCH(" ",{2}{3},1)) & "*", {0}!{4},0))'.format(
            title, rtrn_range, match, row, match_range
        )
    elif dst:
        base_formula = 'INDEX({0}!{1}, MATCH("*(" & {2}{3} & "*", {0}!{4},0))'.format(
            title, rtrn_range, match, row, match_range
        )
    else:
        base_formula = 'INDEX({0}!{1}, MATCH({2}{3} & "*", {0}!{4},0))'.format(
            title, rtrn_range, match, row, match_range
        )

    # if we are looking at weekly stats, blank should be blank, not zero
    if week:
        formula = 'IF(ISBLANK({0}), " ", {0})'.format(base_formula)
    else:
        formula = base_formula

    return "=" + formula


def apply_border(workbook):
    border = Border(
        left=Side(border_style="thin", color="FF000000"),
        right=Side(border_style="thin", color="FF000000"),
    )

    for title in ["QB", "RB", "WR", "TE", "DST"]:
        # select worksheet
        worksheet = workbook[title]
        # find header columns (None = empty cell)
        fields = []
        for cell in worksheet[1]:
            if cell.value is not None:
                fields.append(cell.column_letter)
                # print("field: {} [{}] [idx: {}]".format(cell.value, cell.column, cell.col_idx))

        # add max column (letter) to field
        fields.append(get_column_letter(worksheet.max_column))

        # skip first field
        for i in range(1, len(fields)):
            fmt_range = "{0}1:{1}{2}".format(
                fields[i - 1], chr(ord(fields[i]) - 1), worksheet.max_row
            )
            style_range(worksheet, fmt_range, border=border)


def style_ranges(workbook):
    # define colors for colorscale (from excel)
    red = "F8696B"
    yellow = "FFEB84"
    green = "63BE7B"
    white = "FFFFFF"

    for title in ["QB", "RB", "WR", "TE", "DST"]:
        worksheet = workbook[title]
        # add filter/sort. excel will not automatically do it!
        # filter_range = "{0}:{1}".format('D2', worksheet.max_row)
        # worksheet.auto_filter.ref = filter_range
        # sort_range = "{0}:{1}".format('D3', worksheet.max_row)

        # worksheet.auto_filter.add_sort_condition(sort_range)
        # bigger/positive = green, smaller/negative = red
        green_to_red_headers = [
            "Implied Total",
            "O/U",
            "Run DVOA",
            "Pass DVOA",
            "DVOA",
            "vs. WR1",
            "vs. WR2",
            "O-Line",
            "Snap%",
            "Rush ATTs",
            "Targets",
            "Recepts",
            "vs. TE",
            "D-Line Sack%",
            "Ave PPG",
            "Rushing Yards",
            "DYAR",
            "QBR",
            "Def Yds/Att",
            "Def Comp%",
            "Def TD%",
        ]
        green_to_red_rule = ColorScaleRule(
            start_type="min",
            start_color=red,
            mid_type="percentile",
            mid_value=50,
            mid_color=yellow,
            end_type="max",
            end_color=green,
        )
        # bigger/positive = red, smaller/negative = green
        red_to_green_headers = ["Line", "D-Line", "O-Line Sack%", "ECR"]
        red_to_green_rule = ColorScaleRule(
            start_type="min",
            start_color=green,
            mid_type="percentile",
            mid_value=50,
            mid_color=yellow,
            end_type="max",
            end_color=red,
        )
        white_middle_headers = ["+/- Rank"]
        white_middle_rule = ColorScaleRule(
            start_type="min",
            start_color=red,
            mid_type="percentile",
            mid_value=50,
            mid_color=white,
            end_type="max",
            end_color=green,
        )
        # color ranges
        for i in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=2, column=i).value in green_to_red_headers:
                column_letter = get_column_letter(i)
                # color range (green to red)
                cell_rng = "{0}{1}:{0}{2}".format(column_letter, "3", worksheet.max_row)
                # print("[{}] Coloring {} [{} - {}] green_to_red".format(title, worksheet.cell(row=2, column=i).value, worksheet.cell(row=2, column=i), cell_rng))
                workbook[title].conditional_formatting.add(cell_rng, green_to_red_rule)
            elif worksheet.cell(row=2, column=i).value in red_to_green_headers:
                column_letter = get_column_letter(i)
                # color range (red to green)
                cell_rng = "{0}{1}:{0}{2}".format(column_letter, "3", worksheet.max_row)
                # print("[{}] Coloring {} [{} - {}] red_to_green".format(title, worksheet.cell(row=2, column=i).value, worksheet.cell(row=2, column=i), cell_rng))
                workbook[title].conditional_formatting.add(cell_rng, red_to_green_rule)
            elif worksheet.cell(row=2, column=i).value in white_middle_headers:
                column_letter = get_column_letter(i)
                # color range (red to green)
                cell_rng = "{0}{1}:{0}{2}".format(column_letter, "3", worksheet.max_row)
                # print("[{}] Coloring {} [{} - {}] red_to_green".format(title, worksheet.cell(row=2, column=i).value, worksheet.cell(row=2, column=i), cell_rng))
                workbook[title].conditional_formatting.add(cell_rng, white_middle_rule)


def apply_column_widths(workbook):
    # set column widths
    # column_widths = [8, 20, 10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8]

    # for i, column_width in enumerate(column_widths):
    # worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width
    for title in ["QB", "RB", "WR", "TE", "DST"]:
        try:
            worksheet = workbook[title]
        except KeyError as ex:
            print(f"apply_column_widths(): {ex}")
            return

        for i, cell in enumerate(worksheet[2]):
            # print(cell)
            if cell.value == "Name":
                worksheet.column_dimensions[get_column_letter(i + 1)].width = 20
            elif cell.value == "Opp":
                worksheet.column_dimensions[get_column_letter(i + 1)].width = 10
            elif cell.value == "Position":
                worksheet.column_dimensions[get_column_letter(i + 1)].width = 8
            elif cell.value == "FD Salary":
                worksheet.column_dimensions[get_column_letter(i + 1)].width = 9
            elif cell.value == "+/- Rank":
                worksheet.column_dimensions[get_column_letter(i + 1)].width = 5
            else:
                worksheet.column_dimensions[get_column_letter(i + 1)].width = 7.7


def bool_found_player_in_ecr_tab(ws_column, name):
    # loop through cells in column
    for column in ws_column:
        # if cell is empty, continue
        if column.value is None:
            continue

        # if name is found, move along
        if name in column.value:
            return True
    return False


def freeze_header(workbook):
    # freeze header
    for title in ["QB", "RB", "WR", "TE", "DST"]:
        try:
            worksheet = workbook[title]
        except KeyError as ex:
            print(f"freeze_header(): {ex}")
            return
        worksheet.freeze_panes = "{}3".format(get_column_letter(worksheet.max_column))


def order_sheets(workbook):
    # pull indices from QB, RB, WR, TE, DST to be ordered first
    order = [
        workbook.worksheets.index(workbook[i])
        for i in ["QB", "RB", "WR", "TE", "DST"]
        if i in workbook.sheetnames
    ]

    # create set from 0 to len(workbook._sheets)
    # subtract unique values from set and extend list to fill in missing values
    order.extend(list(set(range(len(workbook._sheets))) - set(order)))
    workbook._sheets = [workbook._sheets[i] for i in order]


def check_name_in_ecr(workbook, position, name):
    # get ECR sheet
    if position + "_ECR" not in workbook.sheetnames:
        return False

    ecr_ws = workbook[position + "_ECR"]
    search_col = "C"

    # search ECR sheet for guy
    return bool_found_player_in_ecr_tab(ecr_ws[search_col], name)


def insert_ranks(workbook):
    for position in ["QB", "RB", "WR", "TE", "DST"]:
        # check if workbook exists
        try:
            worksheet = workbook[position]
        except KeyError as ex:
            print(f"insert_ranks(): {ex}")
            continue

        ecr_col = ""
        ecr_data_col = ""
        salary_col = ""
        salary_rank_col = ""
        plus_minus_col = ""
        max_row = worksheet.max_row

        # look through header row and pull header columns
        for col in worksheet[2]:
            if col.value == "ECR":
                ecr_col = col.column_letter
            elif col.value == "ECR Data":
                ecr_data_col = col.column_letter
            elif col.value == "Salary":
                salary_col = col.column_letter
            elif col.value == "Salary Rank":
                salary_rank_col = col.column_letter
            elif col.value == "+/- Rank":
                plus_minus_col = col.column_letter

        # ECR rank
        for cell in worksheet[ecr_col]:
            # skip header rows
            if cell.row <= 2:
                continue
            cell.value = "=RANK(${0}{1}, ${0}3:${0}{2},1)".format(
                ecr_data_col, cell.row, max_row
            )

        # salary rank
        for cell in worksheet[salary_rank_col]:
            # skip header rows
            if cell.row <= 2:
                continue
            cell.value = "=RANK(${0}{1}, ${0}3:${0}{2},0)".format(
                salary_col, cell.row, max_row
            )

        # +/- rank
        for cell in worksheet[plus_minus_col]:
            # skip header rows
            if cell.row <= 2:
                continue
            cell.value = "={0}{1} - {2}{1}".format(salary_rank_col, cell.row, ecr_col)

        # hide data columns
        # print("1: {}".format(ecr_data_col))
        # print("2: {}".format(salary_rank_col))
        # print("3: {}".format(worksheet))
        # print("4: {}".format(worksheet.column_dimensions[ecr_data_col]))
        worksheet.column_dimensions[ecr_data_col].hidden = True
        worksheet.column_dimensions[salary_rank_col].hidden = True


def read_fantasy_draft_csv(filename):
    with open(filename, "r") as file:
        reader = csv.reader(file)

        # store header row (and strip extra spaces)
        headers = [header.lower().strip() for header in next(reader)]
        headers.append("salary_perc")

        dictionary = {}
        for row in reader:
            # remove periods from name
            row[1] = row[1].replace(".", "")
            # remove Jr. and III etc
            row[1] = " ".join(row[1].split(" ")[:2])

            # store salary without $ or ,
            row[5] = row[5][1:].replace(",", "")
            salary_perc = "{0:0.1%}".format(float(row[5]) / 100000)
            row.append(salary_perc)
            dictionary[row[1]] = {key: value for key, value in zip(headers, row)}
        return dictionary
        # # read entire file into memory
        # lines = file.readlines()

        # for i, line in enumerate(lines):
        #     # skip header
        #     if i == 0:
        #         continue
        #
        #     fields = line.rstrip().split(',')
        #
        #     print(fields)
        #     exit()


# def print_fantasy_draft_to_wb(workbook, fdraft_dict):
#     worksheet = workbook.active

#     for key, value in fdraft_dict.items():
#         player = value
#         # print(player)
#         # worksheet.append([player[key] for key in player])
#         # for k, v in player.items():
#         # print("{}: {}".format(k, v))


def main():
    filename = "DKSalaries_NFL_Sunday_week1.csv"
    dest_filename = "sheet.xlsx"

    # create workbook/worksheet
    workbook = Workbook()
    workbook.guess_types = True  # guess types (numbers, floats, etc)
    ws1 = workbook.active
    ws1.title = "DEL"

    # make sources dir if it does not exist
    directory = "sources"
    if not path.exists(directory):
        makedirs(directory)

    # pull positional stats from fantasypros.com
    for position in ["QB", "RB", "WR", "TE", "DST"]:
        fpros_ecr(workbook, position)

    fdraft_csv = "FDraft_week8_full.csv"
    if path.exists(fdraft_csv):
        fdraft_dict = read_fantasy_draft_csv(fdraft_csv)
    else:
        fdraft_dict = None
    #     print_fantasy_draft_to_wb(workbook, fdraft_dict)
    #     workbook.save(filename=dest_filename)

    with open(filename, "r") as file:
        # read entire file into memory
        lines = file.readlines()

        for i, line in enumerate(lines):
            # skip header
            if i == 0:
                continue

            fields = line.rstrip().split(",")

            # check if player has ECR
            position = fields[0]
            name = fields[2]

            # 'fix' name to remove extra stuff like Jr or III (Todd Gurley II for example)
            name = " ".join(name.split(" ")[:2])
            # also remove periods (T.J. Yeldon for example)
            name = name.replace(".", "")

            if position == "DST":
                name = fields[7]

            # if player does not exist, skip
            if check_name_in_ecr(workbook, position, name) is False:
                # print("Could not find {} [{}]".format(name, position))
                continue

            position_tab(workbook, fields, fields[0], fdraft_dict)

    # pull stats from lineups.com
    # get_nfl_receptions(workbook)
    # get_nfl_targets(workbook)
    # get_nfl_snaps(workbook)
    # get_nfl_rush_atts(workbook)
    # get_nfl_def_stats(workbook)
    # pull stats from footballoutsiders.com
    # get_dvoa_rankings(workbook)
    # get_oline_rankings(workbook)
    # get_dline_rankings(workbook)
    # get_qb_stats_outsiders(workbook)
    # pull vegas stats from oneweekseason.com
    get_vegas_ows(workbook)

    # test
    # write_RB_cols(workbook)

    # set conditional formatting ranges
    style_ranges(workbook)

    # apply left/right borders for sections
    apply_border(workbook)

    # inserts ecr/salary ranks and +/-
    insert_ranks(workbook)

    # apply column widths
    apply_column_widths(workbook)

    # freeze header
    freeze_header(workbook)

    # order sheets
    # workbook._sheets =[workbook._sheets[i] for i in myorder]
    order_sheets(workbook)

    # save workbook (.xlsx file)
    workbook.remove(ws1)  # remove blank worksheet
    workbook.save(filename=dest_filename)

    # remove rows without an ECR ranking (likely out or useless))
    # wb_data_only = load_workbook(dest_filename, data_only=True)

    # TODO ryne hangouts
    # 1 Add a column for Salary rank vs weekly ranking to show best value plays
    #   (example + -, x player ranks 7 but is the 15th most expensive, -8)
    # 2 add Defensive Rank vs QB
    # 3 is it possible to bring in RG rankings? (not a big deal)
    # 4 On the Def tab, I think the implied total should be of the team the defense is against, not their own implied total


if __name__ == "__main__":
    main()
